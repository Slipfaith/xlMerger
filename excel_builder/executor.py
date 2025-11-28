import os
import string
from copy import copy
from typing import Callable, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from utils.logger import logger

LogFn = Callable[[str], None]


class ExcelBuilderExecutor:
    """Applies planned operations and saves Excel files."""

    def __init__(self, log_callback: LogFn | None = None):
        self.log_callback = log_callback or logger.info

    def read_sheets(self, path: str, preview: bool = False) -> Dict[str, pd.DataFrame]:
        try:
            if preview:
                sheets = pd.read_excel(path, sheet_name=None, header=None, nrows=30)
            else:
                sheets = pd.read_excel(path, sheet_name=None, header=None)
            return {name: df.fillna("") for name, df in sheets.items()}
        except Exception as exc:  # noqa: BLE001
            self._log_line(f"Не удалось прочитать {path}: {exc}")
            return {}

    def process_file(self, file_info: Dict[str, str], output_root: str, operations: List[Dict]):
        ext = os.path.splitext(file_info["path"])[1].lower()
        if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            self._process_with_openpyxl(file_info, output_root, operations)
        else:
            self._process_with_pandas(file_info, output_root, operations)

    def _process_with_pandas(self, file_info: Dict[str, str], output_root: str, operations: List[Dict]):
        sheets = self.read_sheets(file_info["path"])
        sheets = self._apply_operations(sheets, operations, file_info["path"])
        rel_path = file_info["rel"]
        dest_path = os.path.join(output_root, rel_path)
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        if dest_path.lower().endswith(".xls"):
            dest_path = dest_path + "x"  # normalize to xlsx for writer
        with pd.ExcelWriter(dest_path, engine="xlsxwriter") as writer:
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=False, header=False)

    def _process_with_openpyxl(self, file_info: Dict[str, str], output_root: str, operations: List[Dict]):
        try:
            workbook = load_workbook(file_info["path"])
        except Exception as exc:  # noqa: BLE001
            self._log_line(f"Не удалось открыть {file_info['path']} с сохранением формата: {exc}")
            self._process_with_pandas(file_info, output_root, operations)
            return

        rename_ops = [op for op in operations if op["type"] == "rename_sheet" and self._op_matches(op, file_info["path"])]
        if rename_ops:
            self._rename_sheets_openpyxl(workbook, rename_ops)

        for op in operations:
            if op["type"] == "rename_sheet":
                continue
            if not self._op_matches(op, file_info["path"]):
                continue
            sheet_name = op.get("sheet")
            if not sheet_name or sheet_name not in workbook.sheetnames:
                self._log_line(
                    f"Пропуск операции {op['type']} для файла {file_info['path']}: лист '{sheet_name}' не найден"
                )
                continue
            sheet = workbook[sheet_name]
            if op["type"] == "rename_header":
                self._rename_header_openpyxl(sheet, op)
            elif op["type"] == "fill_cell":
                self._fill_cell_openpyxl(sheet, op)
            elif op["type"] == "clear_column":
                self._clear_column_openpyxl(sheet, op)

        rel_path = file_info["rel"]
        dest_path = os.path.join(output_root, rel_path)
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        if dest_path.lower().endswith(".xls"):
            dest_path = dest_path + "x"
        workbook.save(dest_path)

    # region operation handling
    def _apply_operations(self, sheets: Dict[str, pd.DataFrame], operations: List[Dict], file_path: str):
        rename_ops = [op for op in operations if op["type"] == "rename_sheet" and self._op_matches(op, file_path)]
        if rename_ops:
            sheets = self._rename_sheets(sheets, rename_ops)

        for op in operations:
            if op["type"] == "rename_sheet":
                continue
            if not self._op_matches(op, file_path):
                continue
            sheet_name = op.get("sheet")
            if not sheet_name or sheet_name not in sheets:
                self._log_line(
                    f"Пропуск операции {op['type']} для файла {file_path}: лист '{sheet_name}' не найден"
                )
                continue
            df = sheets[sheet_name]
            if op["type"] == "rename_header":
                df = self._rename_header(df, op)
            elif op["type"] == "fill_cell":
                df = self._fill_cell(df, op)
            elif op["type"] == "clear_column":
                df = self._clear_column(df, op)
            sheets[sheet_name] = df
        return sheets

    def _rename_sheets(self, sheets: Dict[str, pd.DataFrame], operations: List[Dict]):
        result: Dict[str, pd.DataFrame] = {}
        for name, df in sheets.items():
            new_name = name
            for op in operations:
                if op["old"] == name:
                    candidate = op["new"]
                    counter = 2
                    while candidate in result or candidate in sheets:
                        candidate = f"{op['new']}_{counter}"
                        counter += 1
                    new_name = candidate
                    break
            result[new_name] = df
        return result

    def _rename_sheets_openpyxl(self, workbook, operations: List[Dict]):
        existing = set(workbook.sheetnames)
        for op in operations:
            old = op.get("old")
            new = op.get("new")
            if not old or not new or old not in workbook.sheetnames:
                continue
            candidate = new
            counter = 2
            while candidate in existing:
                candidate = f"{new}_{counter}"
                counter += 1
            workbook[old].title = candidate
            existing.add(candidate)

    def _rename_header(self, df: pd.DataFrame, op: Dict):
        header_row = op.get("header_row", 1) - 1
        col_idx = self._find_column(df, op["identifier"], op["mode"], header_row)
        if col_idx is None:
            self._log_line(f"Колонка {op['identifier']} не найдена")
            return df
        self._ensure_size(df, header_row, col_idx)
        df.iat[header_row, col_idx] = op["new"]
        return df

    def _rename_header_openpyxl(self, sheet, op: Dict):
        header_row = op.get("header_row", 1)
        col_idx = self._find_column_openpyxl(sheet, op["identifier"], op["mode"], header_row)
        if col_idx is None:
            self._log_line(f"Колонка {op['identifier']} не найдена")
            return
        self._ensure_cell(sheet, header_row, col_idx)
        sheet.cell(row=header_row, column=col_idx).value = op.get("new")

    def _fill_cell(self, df: pd.DataFrame, op: Dict):
        col_letter = "".join([c for c in op["cell"] if c.isalpha()])
        row_part = "".join([c for c in op["cell"] if c.isdigit()])
        if not col_letter or not row_part:
            self._log_line(f"Неверная ячейка: {op['cell']}")
            return df
        row_idx = int(row_part) - 1
        col_idx = self._column_from_letter(col_letter)
        self._ensure_size(df, row_idx, col_idx)
        if op.get("only_empty") and pd.notna(df.iat[row_idx, col_idx]):
            return df
        df.iat[row_idx, col_idx] = op.get("value")
        return df

    def _fill_cell_openpyxl(self, sheet, op: Dict):
        col_letter = "".join([c for c in op["cell"] if c.isalpha()])
        row_part = "".join([c for c in op["cell"] if c.isdigit()])
        if not col_letter or not row_part:
            self._log_line(f"Неверная ячейка: {op['cell']}")
            return
        row_idx = int(row_part)
        try:
            col_idx = column_index_from_string(col_letter)
        except ValueError:
            self._log_line(f"Неверная ячейка: {op['cell']}")
            return
        self._ensure_cell(sheet, row_idx, col_idx)
        cell = sheet.cell(row=row_idx, column=col_idx)
        if op.get("only_empty") and cell.value not in (None, ""):
            return
        cell.value = op.get("value")

    def _clear_column(self, df: pd.DataFrame, op: Dict):
        header_row = op.get("header_row", 1) - 1
        col_idx = self._find_column(df, op["identifier"], op["mode"], header_row)
        if col_idx is None:
            self._log_line(f"Колонка {op['identifier']} не найдена")
            return df
        self._ensure_size(df, header_row + 1, col_idx)
        df.iloc[header_row + 1 :, col_idx] = ""
        return df

    def _clear_column_openpyxl(self, sheet, op: Dict):
        header_row = op.get("header_row", 1)
        col_idx = self._find_column_openpyxl(sheet, op["identifier"], op["mode"], header_row)
        if col_idx is None:
            self._log_line(f"Колонка {op['identifier']} не найдена")
            return
        self._ensure_cell(sheet, header_row, col_idx)
        clear_format = op.get("clear_format")
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = ""
            if clear_format:
                self._clear_cell_formatting(cell)

    def _ensure_size(self, df: pd.DataFrame, row_idx: int, col_idx: int):
        while row_idx >= len(df):
            df.loc[len(df)] = [None] * len(df.columns)
        while col_idx >= len(df.columns):
            df[len(df.columns)] = None

    def _find_column(self, df: pd.DataFrame, identifier: str, mode: str, header_row: int):
        if mode == "letter":
            return self._column_from_letter(identifier)
        if header_row < len(df):
            headers = df.iloc[header_row].tolist()
            for idx, val in enumerate(headers):
                if str(val) == identifier:
                    return idx
        return None

    def _find_column_openpyxl(self, sheet, identifier: str, mode: str, header_row: int):
        if mode == "letter":
            try:
                return column_index_from_string(identifier)
            except ValueError:
                return None
        if header_row <= sheet.max_row:
            for cell in sheet[header_row]:
                if str(cell.value) == identifier:
                    return cell.column
        return None

    def _column_from_letter(self, letter: str) -> int:
        letter = letter.upper()
        idx = 0
        for char in letter:
            if char in string.ascii_uppercase:
                idx = idx * 26 + (ord(char) - ord("A") + 1)
        return idx - 1

    def _ensure_cell(self, sheet, row_idx: int, col_idx: int):
        while sheet.max_row < row_idx:
            sheet.append([])
        sheet.cell(row=row_idx, column=col_idx)

    def _clear_cell_formatting(self, cell):
        try:
            base_style = cell.parent.parent._named_styles[0]._style
            cell._style = copy(base_style)
        except Exception:  # noqa: BLE001
            cell._style = copy(cell._style)

    def _op_matches(self, op: Dict, file_path: str) -> bool:
        if op.get("scope") == "all":
            return True
        return op.get("scope") == file_path

    def _log_line(self, text: str):
        self.log_callback(text)

    # endregion
