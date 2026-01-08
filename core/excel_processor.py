import os
import hashlib
from openpyxl import load_workbook
import openpyxl.utils as utils
from openpyxl.styles import PatternFill
from copy import copy as copy_style

from utils.logger import Logger

class ExcelProcessor:
    def __init__(
        self, main_excel_path, folder_path, copy_column, selected_sheets,
        sheet_to_header_row, sheet_to_column, file_to_column=None, folder_to_column=None,
        file_to_sheet_map=None, skip_first_row=False, copy_by_row_number=False,
        preserve_formatting=False, logger=None
    ):
        self.main_excel_path = main_excel_path
        self.folder_path = folder_path
        self.copy_column = copy_column
        self.selected_sheets = selected_sheets
        self.sheet_to_header_row = sheet_to_header_row
        self.sheet_to_column = sheet_to_column
        self.file_to_column = file_to_column or {}
        self.folder_to_column = folder_to_column or {}
        self.file_to_sheet_map = file_to_sheet_map or {}
        self.skip_first_row = skip_first_row
        self.copy_by_row_number = copy_by_row_number
        self.preserve_formatting = preserve_formatting

        self.workbook = None
        self.columns = {}
        self.header_row = {}

        self.logger = logger or Logger()
        self.copy_attempts = 0
        self.copy_successes = 0
        self.copy_failures = []

    @staticmethod
    def get_sheet_names(excel_path):
        wb = load_workbook(excel_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    @staticmethod
    def get_sheet_columns(excel_path, sheet_name, header_row_index):
        wb = load_workbook(excel_path, read_only=True)
        sheet = wb[sheet_name]
        cols = [cell.value for cell in sheet[header_row_index + 1]]
        wb.close()
        return cols

    def validate_paths_and_column(self):
        if self.folder_path and not os.path.isdir(self.folder_path):
            self.logger.log_error("Папка перевода не найдена", "", "", self.folder_path)
            raise FileNotFoundError("Указанная папка не существует.")
        if not os.path.exists(self.main_excel_path):
            self.logger.log_error("Файл Excel не найден", "", "", self.main_excel_path)
            raise FileNotFoundError("Указанный файл Excel не существует.")
        if not self.copy_column:
            self.logger.log_error("Не указан столбец для копирования", "", "", "")
            raise ValueError("Укажите столбец для копирования.")

    def copy_data(self, progress_callback=None):
        self.validate_paths_and_column()
        if not self.selected_sheets:
            self.logger.log_error("Не выбраны листы", "", "", "")
            raise ValueError("Выберите хотя бы один лист.")

        self.workbook = load_workbook(self.main_excel_path)
        self.logger.log_info(f"Загружен основной Excel: {self.main_excel_path}")

        for sheet_name in self.selected_sheets:
            header_row_index = self.sheet_to_header_row[sheet_name]
            self.header_row[sheet_name] = header_row_index
            self.columns[sheet_name] = [
                cell.value for cell in self.workbook[sheet_name][header_row_index + 1]
            ]
            self.logger.log_info(f"Обрабатывается лист: {sheet_name}")

        is_file_mapping = bool(self.file_to_column)
        items = self.file_to_column.items() if is_file_mapping else self.folder_to_column.items()
        total_steps = len(self.selected_sheets) * len([c for _, c in items if c]) if items else 1
        progress = 0

        for sheet_name in self.selected_sheets:
            copy_col_index = utils.column_index_from_string(self.sheet_to_column[sheet_name])
            header_row = self.header_row[sheet_name]

            for name, column_name in items:
                if not column_name:
                    continue
                if column_name not in self.columns[sheet_name]:
                    self.logger.log_error(f"Столбец '{column_name}' не найден на листе '{sheet_name}'", "", "", name)
                    raise Exception(
                        f"Столбец '{column_name}' не найден на листе '{sheet_name}' основного файла Excel."
                    )

                col_index = self.columns[sheet_name].index(column_name) + 1
                if is_file_mapping:
                    file_path = os.path.join(self.folder_path, name)
                    self.logger.log_info(f"Копирование из файла: {file_path}, лист: {sheet_name}, столбец: {column_name}")
                    self._copy_from_file(
                        file_path, sheet_name, copy_col_index, header_row, col_index
                    )
                else:
                    lang_folder_path = os.path.join(self.folder_path, name)
                    self.logger.log_info(f"Копирование из папки: {lang_folder_path}, лист: {sheet_name}, столбец: {column_name}")
                    self._copy_from_folder(
                        lang_folder_path, sheet_name, copy_col_index, header_row, col_index
                    )
                progress += 1
                if progress_callback:
                    progress_callback(progress, total_steps)

        base, ext = os.path.splitext(self.main_excel_path)
        output_file = f"{base}_out{ext}"
        self.workbook.save(output_file)
        self.logger.log_info(f"Файл успешно сохранён: {output_file}")
        report_file = self._write_copy_report(output_file)
        self.logger.log_info(
            f"Итог копирования: всего={self.copy_attempts}, "
            f"успешно={self.copy_successes}, ошибок={len(self.copy_failures)}"
        )
        if report_file:
            self.logger.log_info(f"Отчёт сохранён: {report_file}")
        self.logger.save()
        self.workbook.close()
        return output_file

    def _find_matching_sheet(self, lang_wb, main_sheet_name, file_path=None):
        if file_path:
            mapping = self.file_to_sheet_map.get(file_path, {}).get(main_sheet_name)
            if mapping and mapping in lang_wb.sheetnames:
                return mapping
        if main_sheet_name in lang_wb.sheetnames:
            return main_sheet_name
        elif len(lang_wb.sheetnames) == 1:
            return lang_wb.sheetnames[0]
        else:
            sheets = ', '.join(lang_wb.sheetnames)
            self.logger.log_error(f"Не найден лист '{main_sheet_name}'", "", "", sheets)
            raise Exception(
                f"Не найден лист '{main_sheet_name}' в файле перевода. "
                f"В файле листы: {sheets}. "
                f"Переименуйте листы для автоматического сопоставления, либо удалите лишние листы."
            )

    def _copy_from_file(self, file_path, main_sheet_name, copy_col_index, header_row, col_index):
        if os.path.isfile(file_path) and file_path.endswith(('.xlsx', '.xls')):
            lang_wb = load_workbook(file_path)
            target_sheet_name = self._find_matching_sheet(lang_wb, main_sheet_name, file_path)
            lang_sheet = lang_wb[target_sheet_name]
            self._copy_from_sheet(lang_sheet, main_sheet_name, copy_col_index, header_row, col_index)
            lang_wb.close()

    def _copy_from_folder(self, lang_folder_path, main_sheet_name, copy_col_index, header_row, col_index):
        for filename in os.listdir(lang_folder_path):
            file_path = os.path.join(lang_folder_path, filename)
            if os.path.isfile(file_path) and filename.endswith(('.xlsx', '.xls')):
                lang_wb = load_workbook(file_path)
                target_sheet_name = self._find_matching_sheet(lang_wb, main_sheet_name, file_path)
                lang_sheet = lang_wb[target_sheet_name]
                self._copy_from_sheet(lang_sheet, main_sheet_name, copy_col_index, header_row, col_index)
                lang_wb.close()

    def _copy_from_sheet(self, lang_sheet, sheet_name, copy_col_index, header_row, col_index):
        data_start_row = 2 if self.skip_first_row else 1

        for row in range(data_start_row, lang_sheet.max_row + 1):
            source_value = lang_sheet.cell(row=row, column=copy_col_index).value
            if source_value is None or (isinstance(source_value, str) and source_value.strip() == ""):
                continue
            self.copy_attempts += 1
            data_index = row - data_start_row
            target_row = header_row + 2 + data_index
            source_cell = lang_sheet.cell(row=row, column=copy_col_index)
            if self._set_cell(sheet_name, target_row, col_index, source_value, source_cell):
                self.copy_successes += 1
            else:
                self.copy_failures.append(
                    {
                        "sheet": sheet_name,
                        "row": target_row,
                        "col": col_index,
                        "value": source_value,
                    }
                )

    def _set_cell(self, sheet_name, target_row, col_index, value, source_cell=None):
        target_cell = self.workbook[sheet_name].cell(row=target_row, column=col_index)
        def compute_hash(text):
            if text is None:
                text = ""
            return hashlib.sha256(str(text).encode('utf-8')).hexdigest()
        source_hash = compute_hash(value)
        max_attempts = 5
        for attempt in range(max_attempts):
            target_cell.value = value
            if self.preserve_formatting and source_cell is not None:
                target_cell.font = copy_style(source_cell.font)
                target_cell.border = copy_style(source_cell.border)
                target_cell.fill = copy_style(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy_style(source_cell.protection)
                target_cell.alignment = copy_style(source_cell.alignment)
            if value == target_cell.value and compute_hash(target_cell.value) == source_hash:
                # Успешно скопировано — логируем
                self.logger.log_copy(sheet_name, target_row, col_index, value)
                return True
        else:
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            target_cell.fill = fill
            self.logger.log_error(f"Не удалось записать значение после {max_attempts} попыток", "", "", f"{sheet_name}: R{target_row} C{col_index} [{value}]")
        return False

    def _write_copy_report(self, output_file):
        if self.copy_attempts == 0:
            return None
        base, _ = os.path.splitext(output_file)
        report_file = f"{base}_copy_report.txt"
        with open(report_file, "w", encoding="utf-8") as report:
            report.write("Отчёт о копировании\n")
            report.write(f"Всего попыток: {self.copy_attempts}\n")
            report.write(f"Успешно: {self.copy_successes}\n")
            report.write(f"Ошибки: {len(self.copy_failures)}\n")
            if self.copy_failures:
                report.write("\nСписок ошибок:\n")
                for failure in self.copy_failures:
                    report.write(
                        f"{failure['sheet']} R{failure['row']}C{failure['col']} -> {failure['value']!r}\n"
                    )
        return report_file
