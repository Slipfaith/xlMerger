# -*- coding: utf-8 -*-
import os
import hashlib
from typing import List, Dict
from copy import copy as copy_style
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string


def merge_excel_columns(main_file: str, mappings: List[Dict[str, object]], output_file: str | None = None,
                        progress_callback=None) -> str:
    """Merge columns from multiple Excel files into a main workbook.

    Args:
        main_file: Path to the main Excel workbook.
        mappings: A list of mapping dictionaries. Each mapping must contain:
            - ``source``: path to the source Excel file.
            - ``source_columns``: list of column letters in the source file.
            - ``target_sheet``: sheet name in the main workbook.
            - ``target_columns``: list of column letters in the target sheet.
              ``source_columns`` and ``target_columns`` must have the same length.
        output_file: Optional path where the merged workbook will be saved. If
            not provided, ``main_file`` suffixed with ``_merged`` is used.
        progress_callback: Optional callback function(idx, total, mapping) for progress updates.

    Returns:
        Path to the saved workbook.

    Raises:
        FileNotFoundError: if ``main_file`` or any ``source`` file is missing.
        KeyError: if ``target_sheet`` does not exist in ``main_file``.
        ValueError: if the number of source and target columns differ.
    """
    if not os.path.isfile(main_file):
        raise FileNotFoundError(main_file)

    wb_main = load_workbook(main_file)
    total_mappings = len(mappings)

    try:
        for idx, mp in enumerate(mappings):
            src = mp.get("source")
            src_cols = mp.get("source_columns", [])
            tgt_sheet = mp.get("target_sheet")
            tgt_cols = mp.get("target_columns", [])

            if not os.path.isfile(src):
                raise FileNotFoundError(src)
            if len(src_cols) != len(tgt_cols):
                raise ValueError("Source and target columns must match in length")
            if tgt_sheet not in wb_main.sheetnames:
                raise KeyError(tgt_sheet)

            ws_main = wb_main[tgt_sheet]
            wb_src = load_workbook(src, data_only=True)
            ws_src = wb_src.active

            for s_col, t_col in zip(src_cols, tgt_cols):
                s_idx = column_index_from_string(s_col)
                t_idx = column_index_from_string(t_col)
                for row in range(1, ws_src.max_row + 1):
                    source_cell = ws_src.cell(row=row, column=s_idx)
                    if source_cell.value is None:
                        continue
                    target_cell = ws_main.cell(row=row, column=t_idx)
                    _set_cell_with_retry(target_cell, source_cell)
            wb_src.close()

            if progress_callback:
                progress_callback(idx + 1, total_mappings, mp)

        if output_file is None:
            base, ext = os.path.splitext(main_file)
            output_file = base + "_merged" + ext
        wb_main.save(output_file)
        return output_file
    finally:
        wb_main.close()


def _set_cell_with_retry(target_cell, source_cell, max_attempts: int = 5) -> None:
    """Copy the value and style from ``source_cell`` to ``target_cell`` with retries.

    Some environments sporadically fail to write long text values on the first try.
    This helper ensures the full text is written by verifying the value hash after
    each attempt. Failed writes are highlighted in red.
    """

    def compute_hash(text):
        if text is None:
            text = ""
        return hashlib.sha256(str(text).encode("utf-8")).hexdigest()

    value = source_cell.value
    source_hash = compute_hash(value)

    for _ in range(max_attempts):
        target_cell.value = value
        # copy basic formatting to mimic a real copy-paste
        target_cell.font = copy_style(source_cell.font)
        target_cell.border = copy_style(source_cell.border)
        target_cell.fill = copy_style(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy_style(source_cell.protection)
        target_cell.alignment = copy_style(source_cell.alignment)

        if target_cell.value == value and compute_hash(target_cell.value) == source_hash:
            return

    # mark the cell red if we failed to copy value correctly
    target_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

