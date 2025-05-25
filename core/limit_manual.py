# core/limit_manual.py

from openpyxl.styles import PatternFill

def _get_int_value(value):
    try:
        s = str(value).strip()
        return int(s) if s != "" else None
    except Exception:
        return None

def check_limits_manual(sheet, headers, mappings):
    """
    Проверяет лимиты для ручных сопоставлений (по ячейкам).
    Возвращает: report_lines, total_violations
    """
    report_lines = []
    total_violations = 0

    for m in mappings:
        if m[-1] == "cell":
            selected_cells, manual, upper, lower, _ = m
            current_limit = _get_int_value(upper)
            current_lower = _get_int_value(lower)
            for (model_row, col) in selected_cells:
                excel_row = model_row + 2  # +2 — т.к. первая строка — заголовки
                cell_obj = sheet.cell(row=excel_row, column=col + 1)
                cell_text = cell_obj.value
                if cell_text is None:
                    continue
                text_str = str(cell_text)
                text_length = len(text_str)
                violation = False
                detail = ""
                if current_limit is not None and text_length > current_limit:
                    violation = True
                    detail += f"длина = {text_length} (лимит {current_limit})"
                    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                elif current_lower is not None and text_length < current_lower:
                    violation = True
                    detail += f"длина = {text_length} (нижний лимит {current_lower})"
                    fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
                if violation:
                    cell_obj.fill = fill
                    total_violations += 1
                    header = headers[col]
                    report_lines.append(f"Строка {excel_row}, столбец '{header}': {detail}")

    return report_lines, total_violations