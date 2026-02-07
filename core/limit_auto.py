# -*- coding: utf-8 -*-
# core/limit_auto.py

from openpyxl.styles import PatternFill

def _get_int_value(value):
    try:
        s = str(value).strip()
        return int(s) if s != "" else None
    except Exception:
        return None

def check_limits_auto(sheet, headers, mappings):
    """
    Проверяет лимиты для автоматических сопоставлений (по столбцам).
    Возвращает: report_lines, total_violations, output_file_path, workbook
    """
    report_lines = []
    total_violations = 0

    for mapping in mappings:
        if mapping[-1] != "column":
            continue

        try:
            limit_index = headers.index(mapping[0]) + 1
        except ValueError:
            raise ValueError(f"Столбец '{mapping[0]}' не найден.")

        text_indices = []
        for txt in mapping[1]:
            try:
                text_indices.append(headers.index(txt) + 1)
            except ValueError:
                raise ValueError(f"Столбец '{txt}' не найден.")

        manual = mapping[2]
        upper = mapping[3]
        lower = mapping[4]

        for row in sheet.iter_rows(min_row=2, values_only=False):
            row_num = row[0].row
            limit_cell = row[limit_index - 1]

            if manual:
                current_limit = _get_int_value(upper)
                current_lower = _get_int_value(lower)
            else:
                current_limit = _get_int_value(limit_cell.value)
                current_lower = None

            for txt_idx in text_indices:
                text_cell = row[txt_idx - 1]
                cell_text = text_cell.value
                if cell_text is None:
                    continue
                text_str = str(cell_text)
                text_length = len(text_str)
                violation = False
                detail = ""
                if current_limit is not None and text_length > current_limit:
                    violation = True
                    detail += f"длина = {text_length} (лимит {current_limit})"
                if current_lower is not None and text_length < current_lower:
                    violation = True
                    detail += f", длина = {text_length} (нижний лимит {current_lower})"
                if violation:
                    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    text_cell.fill = fill
                    total_violations += 1
                    header = headers[txt_idx - 1]
                    report_lines.append(f"Строка {row_num}, столбец '{header}': {detail}")

    return report_lines, total_violations