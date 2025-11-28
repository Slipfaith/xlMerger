from openpyxl import load_workbook
import os
import subprocess
from typing import Callable, List, Dict, Tuple
import xlsxwriter
from openpyxl.utils import get_column_letter


def _is_lang_column(name: str) -> bool:
    if not name:
        return False
    name = str(name).strip()
    if len(name) > 5 or ' ' in name or '_' in name:
        return False
    return name.isalpha()


def _normalize_color(color) -> str | None:
    if not color:
        return None
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return None
    if len(rgb) == 8:
        return rgb[-6:]
    if len(rgb) == 6:
        return rgb
    return None


def _get_xlsxwriter_format(workbook, cell, cache):
    if cell is None or not cell.has_style:
        return None

    font = cell.font
    fill = cell.fill
    alignment = cell.alignment

    font_color = _normalize_color(font.color)
    fill_color = _normalize_color(fill.fgColor) if getattr(fill, "fill_type", None) == "solid" else None

    key = (
        bool(font.bold),
        bool(font.italic),
        font.underline,
        font.name,
        font.sz,
        font_color,
        fill_color,
        cell.number_format,
        alignment.horizontal,
        alignment.vertical,
        alignment.wrap_text,
    )

    if key in cache:
        return cache[key]

    fmt_args: Dict[str, object] = {}
    if font.bold:
        fmt_args["bold"] = True
    if font.italic:
        fmt_args["italic"] = True
    if font.underline:
        fmt_args["underline"] = font.underline if isinstance(font.underline, str) else True
    if font.name:
        fmt_args["font_name"] = font.name
    if font.sz:
        fmt_args["font_size"] = font.sz
    if font_color:
        fmt_args["font_color"] = font_color
    if fill_color:
        fmt_args["bg_color"] = fill_color
        fmt_args["pattern"] = 1
    if cell.number_format:
        fmt_args["num_format"] = cell.number_format
    if alignment.horizontal:
        fmt_args["align"] = alignment.horizontal
    if alignment.vertical:
        fmt_args["valign"] = alignment.vertical
    if alignment.wrap_text:
        fmt_args["text_wrap"] = True

    fmt = workbook.add_format(fmt_args)
    cache[key] = fmt
    return fmt


def _write_rows(ws, rows: List[List[Tuple[object, object]]], widths: List[float | None], workbook) -> None:
    fmt_cache: Dict[Tuple, object] = {}
    for idx, width in enumerate(widths):
        if width is not None:
            ws.set_column(idx, idx, width)

    for r_idx, row in enumerate(rows):
        for c_idx, (value, cell) in enumerate(row):
            fmt = _get_xlsxwriter_format(workbook, cell, fmt_cache)
            ws.write(r_idx, c_idx, value, fmt)


def _find_last_data_row(sheet, columns: List[int]) -> int:
    """Return the last row index that has a value in the given columns."""
    for row_idx in range(sheet.max_row, 1, -1):
        for col in columns:
            val = sheet.cell(row=row_idx, column=col).value
            if val not in (None, ""):
                return row_idx
    return 1


def _run_excelltru(file_path: str) -> None:
    """Запуск Excelltru.vbs для file_path (только на Windows)."""

    if os.name != "nt":
        return

    vbs_path = os.path.join(os.path.dirname(__file__), "Excelltru.vbs")
    if not os.path.isfile(vbs_path):
        return

    try:
        norm_path = os.path.normpath(file_path)
        # Оборачиваем путь в кавычки — для пробелов и кириллицы
        subprocess.run(
            ["cscript.exe", "//nologo", vbs_path, f'"{norm_path}"'],
            check=True,
            shell=True
        )
    except Exception as e:
        print(f"⚠ Ошибка при запуске Excelltru.vbs: {e}")


def split_excel_by_languages(
    excel_path: str,
    sheet_name: str,
    source_lang: str,
    output_dir: str | None = None,
    target_langs: list[str] | None = None,
    extra_columns: list[str] | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> List[str]:
    """Split Excel into language pairs."""
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]

    header_map: Dict[str, int] = {}
    col_names: Dict[int, str] = {}
    first_row = next(sheet.iter_rows(min_row=1, max_row=1))
    for idx, cell in enumerate(first_row, start=1):
        letter = get_column_letter(idx)
        header_map[letter] = idx
        val = cell.value
        if val not in (None, ""):
            name = str(val)
            header_map[name] = idx
        else:
            name = letter
        col_names[idx] = name

    if source_lang not in header_map:
        wb.close()
        raise ValueError(f"Source column '{source_lang}' not found")

    if output_dir is None:
        output_dir = os.path.dirname(excel_path)

    target_indices = None
    if target_langs:
        missing = [t for t in target_langs if t not in header_map]
        if missing:
            wb.close()
            raise ValueError(f"Target column(s) {', '.join(missing)} not found")
        target_indices = {header_map[t] for t in target_langs}

    targets: List[tuple[str, int]] = []
    source_idx = header_map[source_lang]
    source_header = col_names[source_idx]
    for idx, name in col_names.items():
        if idx == source_idx:
            continue
        if target_indices is not None:
            if idx not in target_indices:
                continue
        else:
            if not _is_lang_column(name):
                continue
        targets.append((name, idx))

    extra_idx: List[int] = []
    extra_headers: List[str] = []
    if extra_columns:
        for col in extra_columns:
            if col in header_map and header_map[col] != source_idx:
                idx = header_map[col]
                extra_idx.append(idx)
                extra_headers.append(col_names[idx])

    created: List[str] = []
    for i, (target_lang, idx) in enumerate(targets, start=1):
        rows: List[List[Tuple[object, object]]] = []
        widths: List[float | None] = []

        headers: List[Tuple[object, object]] = []
        for header in extra_headers:
            ex_idx = header_map[header]
            cell = sheet.cell(row=1, column=ex_idx)
            headers.append((header, cell))
            widths.append(sheet.column_dimensions[get_column_letter(ex_idx)].width)
        cell = sheet.cell(row=1, column=source_idx)
        headers.append((source_header, cell))
        widths.append(sheet.column_dimensions[get_column_letter(source_idx)].width)
        cell = sheet.cell(row=1, column=idx)
        headers.append((target_lang, cell))
        widths.append(sheet.column_dimensions[get_column_letter(idx)].width)
        rows.append(headers)

        last_row = _find_last_data_row(sheet, [*extra_idx, source_idx, idx])
        for row in range(2, last_row + 1):
            row_data: List[Tuple[object, object]] = []
            for ex_idx in extra_idx:
                src_cell = sheet.cell(row=row, column=ex_idx)
                row_data.append((src_cell.value, src_cell))
            src_cell = sheet.cell(row=row, column=source_idx)
            row_data.append((src_cell.value, src_cell))
            tgt_cell = sheet.cell(row=row, column=idx)
            row_data.append((tgt_cell.value, tgt_cell))
            rows.append(row_data)

        base, ext = os.path.splitext(os.path.basename(excel_path))
        out_name = f"{base}_{source_header}-{target_lang}{ext}"
        out_path = os.path.join(output_dir, out_name)

        wb_out = xlsxwriter.Workbook(out_path)
        ws_new = wb_out.add_worksheet(sheet_name)
        _write_rows(ws_new, rows, widths, wb_out)
        wb_out.close()
        _run_excelltru(out_path)
        created.append(out_path)
        if progress_callback:
            progress_callback(i, len(targets), out_name)

    wb.close()
    return created


def split_excel_multiple_sheets(
    excel_path: str,
    sheet_configs: Dict[str, Tuple[str, List[str] | None, List[str] | None]],
    output_dir: str | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> List[str]:
    """Split multiple sheets preserving sheet names."""
    wb = load_workbook(excel_path)

    if output_dir is None:
        output_dir = os.path.dirname(excel_path)

    workbooks: Dict[str, Dict[str, Dict[str, object]]] = {}
    created: List[str] = []
    source_names: set[str] = set()

    for sheet_name, (src, targets, extras) in sheet_configs.items():
        sheet = wb[sheet_name]
        header_map: Dict[str, int] = {}
        col_names: Dict[int, str] = {}
        first_row = next(sheet.iter_rows(min_row=1, max_row=1))
        for idx, cell in enumerate(first_row, start=1):
            letter = get_column_letter(idx)
            header_map[letter] = idx
            val = cell.value
            if val not in (None, ""):
                name = str(val)
                header_map[name] = idx
            else:
                name = letter
            col_names[idx] = name

        if src not in header_map:
            wb.close()
            raise ValueError(f"Source column '{src}' not found in sheet '{sheet_name}'")

        src_idx = header_map[src]
        src_name = col_names[src_idx]
        source_names.add(src_name)

        target_indices = None
        if targets:
            missing = [t for t in targets if t not in header_map]
            if missing:
                wb.close()
                raise ValueError(
                    f"Target column(s) {', '.join(missing)} not found in sheet '{sheet_name}'"
                )
            target_indices = {header_map[t] for t in targets}

        col_targets: List[Tuple[str, int]] = []
        for idx, name in col_names.items():
            if idx == src_idx:
                continue
            if target_indices is not None:
                if idx not in target_indices:
                    continue
            else:
                if not _is_lang_column(name):
                    continue
            col_targets.append((name, idx))

        extra_idx: List[int] = []
        extra_headers: List[str] = []
        if extras:
            for col in extras:
                if col in header_map and header_map[col] != src_idx:
                    i = header_map[col]
                    extra_idx.append(i)
                    extra_headers.append(col_names[i])

        for tgt_name, idx in col_targets:
            tgt_workbook = workbooks.setdefault(tgt_name, {})
            sheet_info = tgt_workbook.setdefault(sheet_name, {"rows": [], "widths": []})

            if not sheet_info["rows"]:
                widths: List[float | None] = []
                headers: List[Tuple[object, object]] = []
                for header in extra_headers:
                    ex_idx = header_map[header]
                    cell = sheet.cell(row=1, column=ex_idx)
                    headers.append((header, cell))
                    widths.append(sheet.column_dimensions[get_column_letter(ex_idx)].width)
                cell = sheet.cell(row=1, column=src_idx)
                headers.append((src_name, cell))
                widths.append(sheet.column_dimensions[get_column_letter(src_idx)].width)
                cell = sheet.cell(row=1, column=idx)
                headers.append((tgt_name, cell))
                widths.append(sheet.column_dimensions[get_column_letter(idx)].width)
                sheet_info["rows"].append(headers)
                sheet_info["widths"] = widths

            last_row = _find_last_data_row(sheet, [*extra_idx, src_idx, idx])
            for row in range(2, last_row + 1):
                row_data: List[Tuple[object, object]] = []
                for ex_idx in extra_idx:
                    src_cell = sheet.cell(row=row, column=ex_idx)
                    row_data.append((src_cell.value, src_cell))
                src_cell = sheet.cell(row=row, column=src_idx)
                row_data.append((src_cell.value, src_cell))
                tgt_cell = sheet.cell(row=row, column=idx)
                row_data.append((tgt_cell.value, tgt_cell))
                sheet_info["rows"].append(row_data)

    base, ext = os.path.splitext(os.path.basename(excel_path))
    sources = source_names

    src_part = next(iter(sources)) if len(sources) == 1 else "src"

    for i, (tgt, sheets) in enumerate(workbooks.items(), start=1):
        out_name = f"{base}_{src_part}-{tgt}{ext}"
        out_path = os.path.join(output_dir, out_name)
        wb_out = xlsxwriter.Workbook(out_path)
        for sheet_name, info in sheets.items():
            ws_new = wb_out.add_worksheet(sheet_name)
            _write_rows(ws_new, info["rows"], info["widths"], wb_out)
        wb_out.close()
        _run_excelltru(out_path)
        created.append(out_path)
        if progress_callback:
            progress_callback(i, len(workbooks), out_name)

    wb.close()
    return created
