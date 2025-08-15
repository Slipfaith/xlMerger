from PySide6.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableView,
    QListWidget,
    QListWidgetItem,
    QComboBox,
    QListView,
    QCheckBox,
    QHeaderView,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QStandardItemModel, QStandardItem, QColor, QBrush
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from itertools import islice

from gui.limits_checker import DraggableHeaderView
from utils.i18n import tr


class SplitMappingDialog(QDialog):
    """Dialog to select source/target columns for one or many sheets."""

    def __init__(self, excel_path: str, sheet_names, parent=None):
        super().__init__(parent)
        if isinstance(sheet_names, str):
            sheet_names = [sheet_names]
        self.excel_path = excel_path
        self.sheet_names = sheet_names
        self.current_sheet = sheet_names[0]

        self.headers_map: dict[str, list[str]] = {}
        self.models: dict[str, QStandardItemModel] = {}
        # columns that contain at least one non-empty value for each sheet
        self.non_empty_cols: dict[str, set[int]] = {}
        # mapping from sheet name to original column indices
        self.col_indices: dict[str, list[int]] = {}
        # Keep track of which sheets already have preview data loaded
        self._loaded: set[str] = set()
        self.configs = {
            name: {"src": None, "targets": set(), "extra": set()}
            for name in sheet_names
        }

        self.source_col: int | None = None
        self.target_cols: set[int] = set()
        self.extra_cols: set[int] = set()

        # Load preview data only for the first sheet. Additional sheets will be
        # loaded lazily when the user switches to them.
        self._load_preview(self.current_sheet)
        self._init_ui()

    def _load_preview(self, sheet_name: str):
        """Load preview data for a single sheet."""
        if sheet_name in self._loaded:
            return

        wb = load_workbook(self.excel_path, read_only=True)
        sheet = wb[sheet_name]
        headers = [
            str(cell.value) if cell.value is not None else ""
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        rows_for_preview = list(
            sheet.iter_rows(min_row=2, max_row=11, values_only=True)
        )
        rows_for_check = list(islice(sheet.iter_rows(min_row=2, values_only=True), 30))

        non_empty = {idx for idx, h in enumerate(headers) if str(h).strip() != ""}
        if rows_for_check:
            for idx, col in enumerate(zip(*rows_for_check)):
                if any(v not in (None, "") for v in col):
                    non_empty.add(idx)

        keep_idx = sorted(non_empty)
        self.col_indices[sheet_name] = keep_idx
        filtered_headers = [headers[i] for i in keep_idx]
        model = QStandardItemModel()
        header_labels = []
        for idx, orig in enumerate(keep_idx):
            header_value = filtered_headers[idx] if filtered_headers[idx] is not None else ""
            letter = get_column_letter(orig + 1)
            # Show column letter on a separate line to make it always visible
            header_labels.append(f"{letter}\n{header_value}")
        model.setHorizontalHeaderLabels(header_labels)
        for row in rows_for_preview:
            filtered_row = [row[i] if i < len(row) else None for i in keep_idx]
            items = [
                QStandardItem(str(cell)) if cell is not None else QStandardItem("")
                for cell in filtered_row
            ]
            model.appendRow(items)

        self.headers_map[sheet_name] = [
            str(h) if h is not None else "" for h in filtered_headers
        ]
        self.models[sheet_name] = model
        self.non_empty_cols[sheet_name] = set(range(len(keep_idx)))
        self._loaded.add(sheet_name)
        wb.close()

        self.model = self.models[self.current_sheet]
        self.headers = self.headers_map[self.current_sheet]

    def _init_ui(self):
        self.setWindowTitle(tr("Настройка разделения"))
        layout = QVBoxLayout(self)

        if len(self.sheet_names) > 1:
            self.sheet_combo = QComboBox()
            self.sheet_combo.addItems(self.sheet_names)
            self.sheet_combo.currentTextChanged.connect(self.switch_sheet)
            layout.addWidget(self.sheet_combo)

            self.apply_all_checkbox = QCheckBox(tr("Применить настройки первого листа ко всем листам"))
            self.apply_all_checkbox.toggled.connect(self._toggle_apply_all)
            layout.addWidget(self.apply_all_checkbox)

        info = QLabel(
            tr("Выбери исходный столбец (синий) и столбцы перевода (зелёный).")
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setEditTriggers(QTableView.NoEditTriggers)
        self.table.setSelectionMode(QTableView.NoSelection)
        self.table.setSelectionBehavior(QTableView.SelectItems)
        self.header_view = DraggableHeaderView(Qt.Horizontal, self.table)
        self.table.setHorizontalHeader(self.header_view)
        # Ensure the column letters remain visible even with long headers
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.header_view.dragSelectionChanged.connect(self.handle_drag)
        self.header_view.rightClicked.connect(self.handle_right_click)
        self.table.horizontalHeader().setStretchLastSection(False)
        layout.addWidget(self.table)

        layout.addWidget(QLabel(tr("Дополнительные столбцы:")))
        self.extra_list = QListWidget()
        self.extra_list.setFlow(QListView.LeftToRight)
        self.extra_list.setWrapping(True)
        self.extra_list.setMaximumHeight(80)
        non_empty = self.non_empty_cols.get(self.current_sheet, set())
        indices = self.col_indices.get(self.current_sheet, [])
        for idx, h in enumerate(self.headers):
            letter = get_column_letter(indices[idx] + 1) if idx < len(indices) else ""
            item = QListWidgetItem(f"{letter}: {h}")
            item.setData(Qt.UserRole, idx)
            if idx not in non_empty:
                item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
                item.setForeground(QBrush(QColor("gray")))
            item.setCheckState(Qt.Unchecked)
            self.extra_list.addItem(item)
        self.extra_list.itemChanged.connect(self.update_label)
        layout.addWidget(self.extra_list)

        self.count_label = QLabel()
        layout.addWidget(self.count_label)

        self.current_label = QLabel()
        self.current_label.setWordWrap(True)
        layout.addWidget(self.current_label)

        btn_layout = QHBoxLayout()
        ok_btn = QPushButton(tr("Готово"))
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton(tr("Отмена"))
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.update_colors()
        self.update_label()
        self.resize(700, 500)
        self.setFixedSize(self.size())

    def _save_current(self):
        cfg = self.configs[self.current_sheet]
        cfg["src"] = self.source_col
        cfg["targets"] = set(self.target_cols)
        cfg["extra"] = set(self.extra_cols)

    def _toggle_apply_all(self, checked: bool):
        if hasattr(self, "sheet_combo"):
            self.sheet_combo.setEnabled(not checked)
            if checked:
                self.sheet_combo.setCurrentIndex(0)

    def switch_sheet(self, name: str):
        self._save_current()
        self.current_sheet = name
        if name not in self._loaded:
            self._load_preview(name)
        self.model = self.models[name]
        self.headers = self.headers_map[name]
        self.table.setModel(self.model)
        cfg = self.configs[name]
        self.source_col = cfg["src"]
        self.target_cols = set(cfg["targets"])
        self.extra_cols = set(cfg["extra"])
        self._rebuild_extra_list()
        self.update_colors()
        self.update_label()

    def _rebuild_extra_list(self):
        self.extra_list.blockSignals(True)
        self.extra_list.clear()
        non_empty = self.non_empty_cols.get(self.current_sheet, set())
        excluded = {self.source_col} | set(self.target_cols)
        indices = self.col_indices.get(self.current_sheet, [])
        new_extra = set()
        for idx, h in enumerate(self.headers):
            if idx in excluded:
                continue
            letter = get_column_letter(indices[idx] + 1) if idx < len(indices) else ""
            item = QListWidgetItem(f"{letter}: {h}")
            item.setData(Qt.UserRole, idx)
            if idx not in non_empty:
                item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
                item.setForeground(QBrush(QColor("gray")))
            if idx in self.extra_cols:
                item.setCheckState(Qt.Checked)
                new_extra.add(idx)
            else:
                item.setCheckState(Qt.Unchecked)
            self.extra_list.addItem(item)
        self.extra_cols.intersection_update(new_extra)
        self.extra_list.blockSignals(False)

    def handle_drag(self, selection: set[int]):
        start = getattr(self.header_view, "_drag_start", None)
        allowed = self.non_empty_cols.get(self.current_sheet, set())
        if self.source_col is None:
            if start is None:
                return
            if start not in allowed:
                return
            self.source_col = start
            sel = {c for c in selection if c in allowed and c != self.source_col}
            self.target_cols.update(sel)
        else:
            sel = set(selection)
            if self.source_col in sel:
                sel.remove(self.source_col)
            filtered = {c for c in sel if c in allowed}
            self.target_cols.update(filtered)
        self.update_colors()
        self.update_label()

    def handle_right_click(self, index: int):
        if index == self.source_col:
            self.source_col = None
            self.target_cols.clear()
        elif index in self.target_cols:
            self.target_cols.remove(index)
        else:
            return
        self.update_colors()
        self.update_label()

    def update_colors(self):
        for row in range(self.model.rowCount()):
            for col in range(self.model.columnCount()):
                idx = self.model.index(row, col)
                self.model.setData(idx, QBrush(QColor("white")), Qt.BackgroundRole)
        if self.source_col is not None:
            for row in range(self.model.rowCount()):
                idx = self.model.index(row, self.source_col)
                self.model.setData(idx, QBrush(QColor("#a2cffe")), Qt.BackgroundRole)
        for col in self.target_cols:
            for row in range(self.model.rowCount()):
                idx = self.model.index(row, col)
                self.model.setData(idx, QBrush(QColor("#b6fcb6")), Qt.BackgroundRole)
        for col in self.extra_cols:
            for row in range(self.model.rowCount()):
                idx = self.model.index(row, col)
                self.model.setData(idx, QBrush(QColor("#fff2a8")), Qt.BackgroundRole)

    def _collect_extras(self):
        self.extra_cols.clear()
        for i in range(self.extra_list.count()):
            item = self.extra_list.item(i)
            if not item.flags() & Qt.ItemIsEnabled:
                continue
            if item.checkState() == Qt.Checked:
                col_idx = item.data(Qt.UserRole)
                if col_idx not in {self.source_col} | set(self.target_cols):
                    self.extra_cols.add(col_idx)

    def update_label(self):
        # first collect states from the current list
        self._collect_extras()
        # rebuild list to reflect current source/target selections
        self._rebuild_extra_list()
        if self.source_col is None:
            txt = f"{tr('Источник')}: —\n{tr('Цели')}: —\n{tr('Доп')}: —"
        else:
            indices = self.col_indices.get(self.current_sheet, [])
            src_letter = get_column_letter(indices[self.source_col] + 1) if self.source_col < len(indices) else ""
            src = f"{src_letter}: {self.headers[self.source_col]}"
            tgts = (
                [
                    f"{get_column_letter(indices[c] + 1) if c < len(indices) else ''}: {self.headers[c]}"
                    for c in sorted(self.target_cols)
                ]
                if self.target_cols
                else ["—"]
            )
            extras = (
                [
                    f"{get_column_letter(indices[c] + 1) if c < len(indices) else ''}: {self.headers[c]}"
                    for c in sorted(self.extra_cols)
                ]
                if self.extra_cols
                else ["—"]
            )
            txt = (
                f"{tr('Источник')}: {src}\n"
                f"{tr('Цели')}: {', '.join(tgts)}\n"
                f"{tr('Доп')}: {', '.join(extras)}"
            )
        self.count_label.setText(
            tr("Выбрано таргетов: {n}").format(n=len(self.target_cols))
        )
        self.current_label.setText(tr("Текущая настройка:\n{txt}").format(txt=txt))

    def get_selection(self):
        self._save_current()
        if hasattr(self, "apply_all_checkbox") and self.apply_all_checkbox.isChecked():
            base_cfg = self.configs[self.sheet_names[0]]
            if base_cfg["src"] is None:
                return {}
            result = {}
            for sheet in self.sheet_names:
                if sheet not in self._loaded:
                    self._load_preview(sheet)
                headers = self.headers_map[sheet]
                src_idx = base_cfg["src"]
                if src_idx >= len(headers):
                    continue
                src = headers[src_idx]
                tgts = [headers[c] for c in sorted(base_cfg["targets"]) if c < len(headers)]
                extras = [headers[c] for c in sorted(base_cfg["extra"]) if c < len(headers)]
                result[sheet] = (src, tgts, extras)
            return result

        result = {}
        for sheet, cfg in self.configs.items():
            if cfg["src"] is None:
                continue
            headers = self.headers_map[sheet]
            src = headers[cfg["src"]]
            tgts = [headers[c] for c in sorted(cfg["targets"])]
            extras = [headers[c] for c in sorted(cfg["extra"])]
            result[sheet] = (src, tgts, extras)
        return result
