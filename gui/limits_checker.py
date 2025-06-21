import sys
import os
from PySide6.QtWidgets import (
    QApplication, QWidget, QDialog, QVBoxLayout, QHBoxLayout, QTableView, QLabel,
    QLineEdit, QPushButton, QGroupBox, QListWidget, QMenu, QRadioButton,
    QStackedWidget, QTextEdit, QMessageBox, QHeaderView
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QStandardItemModel, QStandardItem, QColor, QBrush
from openpyxl import load_workbook

from core.drag_drop import DragDropLineEdit
from core.limit_auto import check_limits_auto
from core.limit_manual import check_limits_manual

def _get_int_value(value):
    try:
        s = str(value).strip()
        return int(s) if s != "" else None
    except Exception:
        return None

# --- DRAGGABLE HEADER FOR DRAG-SELECT ---
class DraggableHeaderView(QHeaderView):
    dragSelectionChanged = Signal(set)  # set индексов колонок
    rightClicked = Signal(int)  # индекс колонки

    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setSectionsClickable(True)
        self._dragging = False
        self._drag_start = None
        self._drag_current = None

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            index = self.logicalIndexAt(event.position().toPoint())
            if index >= 0:
                self._drag_start = index
                self._drag_current = index
                self._dragging = True
                self.dragSelectionChanged.emit({index})
        elif event.button() == Qt.RightButton:
            index = self.logicalIndexAt(event.position().toPoint())
            if index >= 0:
                self.rightClicked.emit(index)
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._dragging:
            index = self.logicalIndexAt(event.position().toPoint())
            if index >= 0 and index != self._drag_current:
                self._drag_current = index
                start = min(self._drag_start, self._drag_current)
                end = max(self._drag_start, self._drag_current)
                selection = set(range(start, end + 1))
                self.dragSelectionChanged.emit(selection)
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._dragging = False
        super().mouseReleaseEvent(event)

# --- FILE SELECTION PAGE (DRAG&DROP, 2x CLICK) ---
class FileSelectionPage(QWidget):
    file_selected = Signal(str)
    mapping_clicked = Signal()
    next_clicked = Signal()

    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        self.file_input = DragDropLineEdit(mode='file')
        self.file_input.setPlaceholderText("Перетащи Excel-файл сюда или дважды кликни для выбора...")
        self.file_input.fileSelected.connect(self.file_selected_handler)
        layout.addWidget(QLabel("Файл Excel:"))
        layout.addWidget(self.file_input)
        self.sheet_label = QLabel("Лист: (не выбран)")
        layout.addWidget(self.sheet_label)
        self.mapping_btn = QPushButton("Проверить лимиты")
        self.mapping_btn.clicked.connect(self.mapping_clicked.emit)
        layout.addWidget(self.mapping_btn)
        self.next_btn = QPushButton("Далее")
        self.next_btn.clicked.connect(self.next_clicked.emit)
        layout.addWidget(self.next_btn)
        self._sheetnames = []
        self._current_sheet = ""

    def file_selected_handler(self, file_path):
        self.file_selected.emit(file_path)

    def set_sheets(self, sheets):
        self._sheetnames = sheets
        self._current_sheet = sheets[0] if sheets else ""
        self.sheet_label.setText(f"Лист: {self._current_sheet}")

    def current_sheet(self):
        return self._current_sheet

# --- MAPPING DIALOG ---
class MappingDialog(QDialog):
    def __init__(self, model: QStandardItemModel, headers: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Сопоставление лимитов")
        self.resize(800, 600)
        self.headers = headers
        self.model = model
        self.mappings = []

        self.mode_auto = True
        self.current_limit_col = None
        self.current_text_cols = set()
        self.manual_selected = set()
        self.saved_manual_cells = set()
        self.saved_auto_cells = set()
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)

        # Переключатель режимов
        mode_layout = QHBoxLayout()
        self.auto_radio = QRadioButton("Автоматический режим")
        self.manual_radio = QRadioButton("Ручной режим")
        self.auto_radio.setChecked(True)
        self.auto_radio.toggled.connect(self.switch_mode)
        mode_layout.addWidget(self.auto_radio)
        mode_layout.addWidget(self.manual_radio)
        main_layout.addLayout(mode_layout)

        # Инструкции
        self.info_label = QLabel(
            "Автоматический: выбери столбец лимитов (синий), затем перетяни мышкой по заголовкам — выделишь текстовые столбцы (зелёный).\n"
            "Ручной: выдели любые ячейки мышкой, введи лимиты вручную."
        )
        self.info_label.setWordWrap(True)
        main_layout.addWidget(self.info_label)

        # Таблица предпросмотра
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setEditTriggers(QTableView.NoEditTriggers)
        self.table.setSelectionMode(QTableView.NoSelection)
        self.table.setSelectionBehavior(QTableView.SelectItems)
        self.header_view = DraggableHeaderView(Qt.Horizontal, self.table)
        self.table.setHorizontalHeader(self.header_view)
        self.header_view.dragSelectionChanged.connect(self.handle_drag_selection)
        self.table.selectionModel().selectionChanged.connect(self.on_selection_changed)
        main_layout.addWidget(self.table)

        # Поля лимитов (только для ручного режима)
        self.manual_group = QGroupBox("Ручной ввод лимитов")
        manual_layout = QHBoxLayout()
        self.upper_limit_edit = QLineEdit()
        self.upper_limit_edit.setPlaceholderText("Верхний лимит")
        self.upper_limit_edit.setFixedWidth(100)
        self.lower_limit_edit = QLineEdit()
        self.lower_limit_edit.setPlaceholderText("Нижний лимит")
        self.lower_limit_edit.setFixedWidth(100)
        manual_layout.addWidget(self.upper_limit_edit)
        manual_layout.addWidget(self.lower_limit_edit)
        self.manual_group.setLayout(manual_layout)
        self.manual_group.setVisible(False)
        main_layout.addWidget(self.manual_group)

        # Текущий выбор
        self.current_label = QLabel("Текущая настройка: —")
        self.current_label.setWordWrap(True)
        main_layout.addWidget(self.current_label)

        # Кнопки
        btn_layout = QHBoxLayout()
        self.save_btn = QPushButton("Подтвердить")
        self.save_btn.clicked.connect(self.save_mapping)
        self.clear_btn = QPushButton("Очистить выбор")
        self.clear_btn.clicked.connect(self.clear_selection)
        btn_layout.addWidget(self.save_btn)
        btn_layout.addWidget(self.clear_btn)
        main_layout.addLayout(btn_layout)

        # Список сопоставлений
        main_layout.addWidget(QLabel("Сохранённые сопоставления:"))
        self.mapping_list = QListWidget()
        self.mapping_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.mapping_list.customContextMenuRequested.connect(self.show_context_menu)
        main_layout.addWidget(self.mapping_list)

        # Готово/Отмена
        bottom_layout = QHBoxLayout()
        done_btn = QPushButton("Готово")
        done_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        bottom_layout.addWidget(done_btn)
        bottom_layout.addWidget(cancel_btn)
        main_layout.addLayout(bottom_layout)

        self.setLayout(main_layout)

    def switch_mode(self):
        self.mode_auto = self.auto_radio.isChecked()
        self.manual_group.setVisible(not self.mode_auto)
        self.table.clearSelection()
        self.current_limit_col = None
        self.current_text_cols.clear()
        self.manual_selected.clear()
        if self.mode_auto:
            self.header_view.setEnabled(True)
            self.table.setSelectionMode(QTableView.NoSelection)
            self.table.setAutoScroll(False)
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        else:
            self.header_view.setEnabled(False)
            self.table.setSelectionMode(QTableView.ExtendedSelection)
            self.table.setAutoScroll(True)
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.update_colors()
        self.update_label()

    def handle_drag_selection(self, selection: set):
        if not self.mode_auto:
            return
        if self.current_limit_col is None and len(selection) == 1:
            self.current_limit_col = next(iter(selection))
            self.current_text_cols.clear()
        elif self.current_limit_col is not None:
            sel = set(selection)
            if self.current_limit_col in sel:
                sel.remove(self.current_limit_col)
            self.current_text_cols = sel
        self.update_colors()
        self.update_label()

    def on_selection_changed(self):
        if self.mode_auto:
            return
        self.manual_selected = set(
            (index.row(), index.column())
            for index in self.table.selectionModel().selectedIndexes()
        )
        self.update_colors()
        self.update_label()

    def update_colors(self):
        # Сброс всех цветов
        for row in range(self.model.rowCount()):
            for col in range(self.model.columnCount()):
                idx = self.model.index(row, col)
                self.model.setData(idx, QBrush(QColor("white")), Qt.BackgroundRole)
        # Оранжевый — уже подтверждённые (ручные и авто)
        for row, col in self.saved_manual_cells:
            idx = self.model.index(row, col)
            self.model.setData(idx, QBrush(QColor("#ffe5b4")), Qt.BackgroundRole)
        for row, col in self.saved_auto_cells:
            idx = self.model.index(row, col)
            self.model.setData(idx, QBrush(QColor("#ffe5b4")), Qt.BackgroundRole)
        # Выделение текущего выбора
        if self.mode_auto:
            if self.current_limit_col is not None:
                for row in range(self.model.rowCount()):
                    idx = self.model.index(row, self.current_limit_col)
                    self.model.setData(idx, QBrush(QColor("#a2cffe")), Qt.BackgroundRole)
            for col in self.current_text_cols:
                for row in range(self.model.rowCount()):
                    idx = self.model.index(row, col)
                    self.model.setData(idx, QBrush(QColor("#b6fcb6")), Qt.BackgroundRole)
        else:
            for row, col in self.manual_selected:
                idx = self.model.index(row, col)
                self.model.setData(idx, QBrush(QColor("#b6fcb6")), Qt.BackgroundRole)

    def update_label(self):
        if self.mode_auto:
            if self.current_limit_col is None:
                txt = "Лимит: —; Тексты: —"
            else:
                lim = self.headers[self.current_limit_col]
                texts = [self.headers[c] for c in sorted(self.current_text_cols)] if self.current_text_cols else ['—']
                txt = f"Лимит: {lim}; Тексты: {', '.join(texts)}"
        else:
            if not self.manual_selected:
                txt = "Ячеек выбрано: —"
            else:
                cells = [f"({r + 2}, {self.headers[c]})" for r, c in sorted(self.manual_selected)]
                up = self.upper_limit_edit.text() or '—'
                low = self.lower_limit_edit.text() or '—'
                txt = f"Ячейки: {', '.join(cells)}; Верхний: {up}; Нижний: {low}"
        self.current_label.setText(f"Текущая настройка: {txt}")

    def clear_selection(self):
        self.table.clearSelection()
        self.current_limit_col = None
        self.current_text_cols.clear()
        self.manual_selected.clear()
        self.upper_limit_edit.clear()
        self.lower_limit_edit.clear()
        self.update_colors()
        self.update_label()

    def save_mapping(self):
        if self.mode_auto:
            if self.current_limit_col is None or not self.current_text_cols:
                QMessageBox.critical(self, "Ошибка", "Выберите лимитный столбец и хотя бы одну колонку с текстом.")
                return
            mapping = (
                self.headers[self.current_limit_col],
                [self.headers[c] for c in sorted(self.current_text_cols)],
                False,
                None,
                None,
                "column"
            )
            txt = f"Лимит: {mapping[0]} -> {', '.join(mapping[1])}"
            # --- Оранжевое выделение подтверждённых ячеек (лимит + тексты)
            for row in range(self.model.rowCount()):
                self.saved_auto_cells.add((row, self.current_limit_col))
                for col in self.current_text_cols:
                    self.saved_auto_cells.add((row, col))
        else:
            if not self.manual_selected:
                QMessageBox.critical(self, "Ошибка", "Выделите ячейки для проверки.")
                return
            upper = _get_int_value(self.upper_limit_edit.text())
            lower = _get_int_value(self.lower_limit_edit.text())
            mapping = (
                list(self.manual_selected),
                True,
                upper,
                lower,
                "cell"
            )
            cells = ', '.join([f"({r + 2},{self.headers[c]})" for r, c in mapping[0]])
            txt = f"Ячейки: {cells}; Верхний={upper if upper is not None else '—'}, Нижний={lower if lower is not None else '—'}"
            self.saved_manual_cells.update(self.manual_selected)
        self.mappings.append(mapping)
        self.mapping_list.addItem(txt)
        self.clear_selection()

    def show_context_menu(self, pos):
        item = self.mapping_list.itemAt(pos)
        if item:
            menu = QMenu()
            delete_action = menu.addAction("Удалить")
            action = menu.exec(self.mapping_list.mapToGlobal(pos))
            row = self.mapping_list.row(item)
            if action == delete_action:
                self.mapping_list.takeItem(row)
                if 0 <= row < len(self.mappings):
                    del self.mappings[row]
                # Оставляем saved_manual_cells/saved_auto_cells — чтобы покраска оставалась

    def get_mappings(self):
        return self.mappings

# --- LIMITS CHECKER MAIN ---
class LimitsChecker(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Проверка лимитов")
        self.resize(800, 600)
        self.stack = QStackedWidget()
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.stack)
        self.setLayout(main_layout)
        self.selected_file = ""
        self.workbook = None
        self.sheet = None
        self.sheet_name = ""
        self.headers = []
        self.mappings = []
        self.report_text = ""

        self.file_page = FileSelectionPage()
        self.stack.addWidget(self.file_page)
        self.file_page.file_selected.connect(self.update_sheet_list)
        self.file_page.mapping_clicked.connect(self.open_mapping_dialog)
        self.file_page.next_clicked.connect(self.goto_results_page)

    def update_sheet_list(self, file_path):
        try:
            wb = load_workbook(file_path, read_only=True)
            self.file_page.set_sheets(wb.sheetnames)
            self.selected_file = file_path
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить листы: {e}")

    def open_mapping_dialog(self):
        if not self.selected_file:
            QMessageBox.critical(self, "Ошибка", "Выберите файл Excel.")
            return
        self.sheet_name = self.file_page.current_sheet()
        try:
            self.workbook = load_workbook(self.selected_file)
            self.sheet = self.workbook[self.sheet_name]
            self.headers = [str(cell.value) if cell.value is not None else ""
                            for cell in next(self.sheet.iter_rows(min_row=1, max_row=1))]
            if not any(self.headers):
                QMessageBox.critical(self, "Ошибка", "Не найдены заголовки в первой строке.")
                return
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при открытии файла: {e}")
            return
        model = QStandardItemModel()
        model.setHorizontalHeaderLabels(self.headers)
        row_count = 10
        rows = list(self.sheet.iter_rows(min_row=2, max_row=row_count + 1, values_only=True))
        for row in rows:
            items = [QStandardItem(str(cell)) if cell is not None else QStandardItem("")
                     for cell in row]
            model.appendRow(items)
        dialog = MappingDialog(model, self.headers, self)
        if dialog.exec() == QDialog.Accepted:
            self.mappings = dialog.get_mappings()

    def goto_results_page(self):
        if not self.mappings:
            QMessageBox.critical(self, "Ошибка", "Сначала создайте сопоставления через кнопку 'Лимиты'.")
            return
        self.run_limit_check()

    def run_limit_check(self):
        if not self.mappings:
            QMessageBox.critical(self, "Ошибка", "Не заданы сопоставления лимитов.")
            return
        report_lines = []
        total_violations = 0

        try:
            auto_lines, auto_violations = check_limits_auto(self.sheet, self.headers, self.mappings)
            report_lines.extend(auto_lines)
            total_violations += auto_violations
        except ValueError as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            return

        manual_lines, manual_violations = check_limits_manual(self.sheet, self.headers, self.mappings)
        report_lines.extend(manual_lines)
        total_violations += manual_violations

        base, ext = os.path.splitext(self.selected_file)
        output_file = f"{base}_checked{ext}"
        try:
            self.workbook.save(output_file)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {e}")
            return
        report = "Проверка лимитов завершена.\n"
        report += f"Всего нарушений: {total_violations}\n\n"
        if report_lines:
            report += "Детали:\n" + "\n".join(report_lines)
        else:
            report += "Нарушений не обнаружено."
        self.report_text = report
        self.results_page = self.create_results_page(output_file)
        self.stack.addWidget(self.results_page)
        self.stack.setCurrentWidget(self.results_page)

    def create_results_page(self, output_file):
        page = QWidget()
        layout = QVBoxLayout()
        report_label = QLabel("Результаты проверки:")
        layout.addWidget(report_label)
        self.report_text_edit = QTextEdit()
        self.report_text_edit.setReadOnly(True)
        self.report_text_edit.setText(self.report_text)
        layout.addWidget(self.report_text_edit)
        file_info = QLabel(f"Изменённый файл сохранён как:\n{output_file}")
        layout.addWidget(file_info)
        btn_layout = QHBoxLayout()
        back_btn = QPushButton("Вернуться к сопоставлению")
        back_btn.clicked.connect(self.go_back_to_file_page)
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(back_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
        page.setLayout(layout)
        return page

    def go_back_to_file_page(self):
        self.stack.setCurrentWidget(self.file_page)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = LimitsChecker()
    window.show()
    sys.exit(app.exec())