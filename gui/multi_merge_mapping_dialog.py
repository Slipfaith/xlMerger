# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional

from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QVBoxLayout,
    QWidget,
)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.i18n import tr
from .merge_mapping_dialog import MergeMappingDialog


def _detect_header_row(ws, max_rows: int = 5) -> int:
    """Return the most likely header row index (1-based)."""

    best_row = 1
    best_score = -1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        values = [cell for cell in row if cell not in (None, "")]
        text_score = sum(1 for cell in values if isinstance(cell, str))
        score = text_score * 2 + len(values)
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row


def _normalize_headers(header_values) -> List[str]:
    """Convert row values to strings and drop trailing empty headers."""
    headers = [str(val).strip() if val is not None else "" for val in header_values]
    while headers and not headers[-1]:
        headers.pop()
    return headers


def _read_structure(path: str) -> Dict[str, Dict[str, object]]:
    """Collect sheet headers and guessed header rows for similarity checks."""

    workbook = load_workbook(path, read_only=True)
    structure: Dict[str, Dict[str, object]] = {}
    try:
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            header_row = _detect_header_row(ws)
            header_values = next(
                ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
                (),
            )
            headers = _normalize_headers(header_values)
            structure[sheet_name] = {
                "header_row": header_row,
                "headers": headers,
            }
    finally:
        workbook.close()
    return structure


def _score_match(target_struct: Dict[str, Dict[str, object]], source_struct: Dict[str, Dict[str, object]],
                 target_name: str, source_name: str) -> float:
    """Heuristic similarity score between two Excel files."""

    score = 0.0
    base_target = os.path.splitext(os.path.basename(target_name))[0].lower()
    base_source = os.path.splitext(os.path.basename(source_name))[0].lower()
    if base_target and base_target == base_source:
        score += 5
    elif base_target and base_target in base_source:
        score += 3

    shared_sheets = set(target_struct.keys()) & set(source_struct.keys())
    for sheet in shared_sheets:
        score += 2
        tgt_headers = {h.lower() for h in target_struct[sheet].get("headers", []) if h}
        src_headers = {h.lower() for h in source_struct[sheet].get("headers", []) if h}
        if tgt_headers:
            overlap = len(tgt_headers & src_headers)
            score += overlap / max(len(tgt_headers), 1)

    if not shared_sheets and target_struct and source_struct:
        # fallback: compare first sheets by headers
        tgt_sheet = next(iter(target_struct.values()))
        src_sheet = next(iter(source_struct.values()))
        tgt_headers = {h.lower() for h in tgt_sheet.get("headers", []) if h}
        src_headers = {h.lower() for h in src_sheet.get("headers", []) if h}
        overlap = len(tgt_headers & src_headers)
        score += overlap / max(len(tgt_headers) or 1, 1)
    return score


def _auto_column_mapping(target_path: str, source_path: str,
                         target_structures: Dict[str, Dict[str, object]],
                         source_structures: Dict[str, Dict[str, object]]) -> List[Dict[str, object]]:
    """Build column mappings by matching equal headers across sheets."""

    mappings: List[Dict[str, object]] = []
    tgt_struct = target_structures.get(target_path, {})
    src_struct = source_structures.get(source_path, {})

    for sheet_name, src_info in src_struct.items():
        target_sheet = sheet_name if sheet_name in tgt_struct else None
        if not target_sheet and tgt_struct:
            # pick sheet with maximum header overlap
            best_overlap = 0
            src_headers_set = {h.lower() for h in src_info.get("headers", []) if h}
            for candidate, tgt_info in tgt_struct.items():
                tgt_headers = {h.lower() for h in tgt_info.get("headers", []) if h}
                overlap = len(tgt_headers & src_headers_set)
                if overlap > best_overlap:
                    best_overlap = overlap
                    target_sheet = candidate

        if not target_sheet:
            continue

        tgt_headers = tgt_struct.get(target_sheet, {}).get("headers", [])
        src_headers = src_info.get("headers", [])
        header_to_target_idx: Dict[str, int] = {}
        for tgt_idx, tgt_header in enumerate(tgt_headers):
            norm = str(tgt_header).strip().lower()
            if norm and norm not in header_to_target_idx:
                header_to_target_idx[norm] = tgt_idx

        pairs: List[tuple[int, int]] = []
        for idx, header in enumerate(src_headers):
            norm = str(header).strip().lower()
            if not norm:
                continue
            tgt_idx = header_to_target_idx.get(norm)
            if tgt_idx is not None:
                pairs.append((idx, tgt_idx))

        if pairs:
            mappings.append({
                "source": source_path,
                "source_columns": [get_column_letter(i + 1) for i, _ in pairs],
                "target_sheet": target_sheet,
                "target_columns": [get_column_letter(j + 1) for _, j in pairs]
            })
    return mappings


@dataclass
class MappingRowState:
    target_path: str
    selector: QComboBox
    status_label: QLabel
    settings_button: QPushButton
    mappings: List[Dict[str, object]] = field(default_factory=list)
    selected_sources: List[str] = field(default_factory=list)


class MultiMergeMappingDialog(QDialog):
    def __init__(self, targets: List[str], translations: List[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr("Сопоставление файлов"))
        self.resize(850, 500)

        self.targets = targets
        self.translations = translations
        self.target_structures = {path: _read_structure(path) for path in targets}
        self.translation_structures = {path: _read_structure(path) for path in translations}
        self.rows: List[MappingRowState] = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        header = QLabel(tr("Проверь сопоставления файлов перевода с целевыми Excel."))
        header.setWordWrap(True)
        layout.addWidget(header)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setSpacing(12)

        for target in self.targets:
            group = QGroupBox(os.path.basename(target))
            group_layout = QGridLayout()

            group_layout.addWidget(QLabel(tr("Файл перевода")), 0, 0)
            selector = QComboBox()
            selector.addItem(tr("Не выбрано"), None)
            for src in self.translations:
                selector.addItem(os.path.basename(src), src)
            group_layout.addWidget(selector, 0, 1)

            status_label = QLabel()
            status_label.setWordWrap(True)
            group_layout.addWidget(status_label, 1, 0, 1, 2)

            settings_btn = QPushButton(tr("Настроить столбцы"))
            group_layout.addWidget(settings_btn, 0, 2)

            group.setLayout(group_layout)
            container_layout.addWidget(group)

            row = MappingRowState(target, selector, status_label, settings_btn)
            selector.currentIndexChanged.connect(lambda _, r=row: self._on_selection_changed(r))
            settings_btn.clicked.connect(lambda _, r=row: self._open_columns_dialog(r))
            self.rows.append(row)
            self._apply_auto_selection(row)

        container_layout.addStretch()
        scroll.setWidget(container)
        layout.addWidget(scroll)

        buttons = QHBoxLayout()
        buttons.addStretch()
        ok_btn = QPushButton(tr("Готово"))
        ok_btn.clicked.connect(self._on_accept)
        cancel_btn = QPushButton(tr("Отмена"))
        cancel_btn.clicked.connect(self.reject)
        buttons.addWidget(ok_btn)
        buttons.addWidget(cancel_btn)
        layout.addLayout(buttons)

    def _apply_auto_selection(self, row: MappingRowState):
        best_source: Optional[str] = None
        best_score = -1.0
        second_score = -1.0
        for src in self.translations:
            score = _score_match(
                self.target_structures.get(row.target_path, {}),
                self.translation_structures.get(src, {}),
                row.target_path,
                src,
            )
            if score > best_score:
                second_score = best_score
                best_score = score
                best_source = src
            elif score > second_score:
                second_score = score

        if best_source:
            idx = row.selector.findData(best_source)
            if idx >= 0:
                row.selector.setCurrentIndex(idx)
        ambiguity = best_score > 0 and second_score >= best_score - 0.5
        if ambiguity:
            row.status_label.setText(tr("Найдено несколько вариантов. Проверь выбор вручную."))
        elif best_score <= 0:
            row.status_label.setText(tr("Не удалось подобрать перевод автоматически."))
        else:
            row.status_label.setText(tr("Автовыбор выполнен по совпадению листов/заголовков."))

        if best_source:
            row.selected_sources = [best_source]
            row.mappings = _auto_column_mapping(
                row.target_path,
                best_source,
                self.target_structures,
                self.translation_structures,
            )

    def _on_selection_changed(self, row: MappingRowState):
        selected = row.selector.currentData()
        if not selected:
            row.status_label.setText(tr("Выберите файл перевода."))
            row.mappings = []
            row.selected_sources = []
            return

        row.selected_sources = [selected]
        row.mappings = _auto_column_mapping(
            row.target_path,
            selected,
            self.target_structures,
            self.translation_structures,
        )
        if row.mappings:
            row.status_label.setText(tr("Столбцы подобраны по совпадающим заголовкам."))
        else:
            row.status_label.setText(tr("Не удалось сопоставить столбцы автоматически."))

    def _open_columns_dialog(self, row: MappingRowState):
        if not row.selected_sources:
            QMessageBox.warning(self, tr("Предупреждение"), tr("Сначала выбери файл перевода."))
            return

        dialog = MergeMappingDialog(row.target_path, self)
        for src in row.selected_sources:
            dialog.add_row_with_file(src)
        if dialog.exec():
            mappings = dialog.get_mappings()
            if mappings:
                row.mappings = mappings
                row.status_label.setText(tr("Пользовательские сопоставления применены."))

    def _on_accept(self):
        unresolved = [row for row in self.rows if not row.mappings]
        if unresolved:
            names = "\n".join(os.path.basename(r.target_path) for r in unresolved)
            QMessageBox.warning(self, tr("Предупреждение"), tr("Не заданы сопоставления для:\n{names}").format(names=names))
            return
        self.accept()

    def get_tasks(self) -> List[Dict[str, object]]:
        tasks: List[Dict[str, object]] = []
        for row in self.rows:
            if row.mappings:
                tasks.append({
                    "target": row.target_path,
                    "mappings": row.mappings,
                })
        return tasks
