# -*- coding: utf-8 -*-
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QWidget, QFileDialog,
    QLabel, QComboBox, QMessageBox, QScrollArea, QFrame, QGridLayout,
    QLineEdit, QRadioButton, QButtonGroup
)
from PySide6.QtCore import Qt
import os
from .style_system import (
    set_button_shape,
    set_button_variant,
    set_label_role,
)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    get_column_letter = None
try:
    import xlrd
except ImportError:
    xlrd = None


def get_excel_structure(path):
    sheets = {}
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx" and openpyxl:
        wb = openpyxl.load_workbook(path, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(val) if val is not None else "" for val in row]
            sheets[sheet] = headers
        wb.close()
    elif ext == ".xls" and xlrd:
        wb = xlrd.open_workbook(path)
        for sheet in wb.sheet_names():
            ws = wb.sheet_by_name(sheet)
            headers = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]
            sheets[sheet] = headers
    else:
        raise Exception("Unsupported file format or required lib missing")
    return sheets


class MappingCard(QFrame):
    def __init__(self, main_structure, parent=None):
        super().__init__(parent)
        self.main_structure = main_structure
        self.file_path = None
        self.file_structure = {}
        self.setObjectName("mappingCard")

        self.setFrameStyle(QFrame.Box)
        self.setLineWidth(1)

        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        # Заголовок карточки с кнопкой удаления
        header_layout = QHBoxLayout()

        self.title_label = QLabel("Новое сопоставление")
        set_label_role(self.title_label, "heading")
        header_layout.addWidget(self.title_label)

        header_layout.addStretch()

        self.remove_btn = QPushButton("✕")
        self.remove_btn.setFixedSize(25, 25)
        set_button_variant(self.remove_btn, "danger")
        set_button_shape(self.remove_btn, "circleLarge")
        header_layout.addWidget(self.remove_btn)

        layout.addLayout(header_layout)

        # Выбор файла источника
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("Файл:"))

        self.file_btn = QPushButton("Выбрать файл...")
        self.file_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.file_btn)

        layout.addLayout(file_layout)

        # Выбор листов
        sheets_layout = QGridLayout()

        sheets_layout.addWidget(QLabel("Лист источника:"), 0, 0)
        self.source_sheet_combo = QComboBox()
        self.source_sheet_combo.setEnabled(False)
        self.source_sheet_combo.currentTextChanged.connect(self.update_title)
        sheets_layout.addWidget(self.source_sheet_combo, 0, 1)

        sheets_layout.addWidget(QLabel("Лист назначения:"), 0, 2)
        self.target_sheet_combo = QComboBox()
        self.target_sheet_combo.addItems(list(self.main_structure.keys()))
        self.target_sheet_combo.currentTextChanged.connect(self.update_title)
        sheets_layout.addWidget(self.target_sheet_combo, 0, 3)

        layout.addLayout(sheets_layout)

        # Режим сопоставления
        mode_group = QFrame()
        mode_group.setFrameStyle(QFrame.StyledPanel)
        mode_layout = QVBoxLayout(mode_group)

        mode_label = QLabel("Режим сопоставления:")
        set_label_role(mode_label, "heading")
        mode_layout.addWidget(mode_label)

        self.mode_group = QButtonGroup()

        self.letter_mode = QRadioButton("По буквенным обозначениям (A, B, C...)")
        self.letter_mode.setChecked(True)
        self.letter_mode.toggled.connect(self.on_mode_changed)
        self.mode_group.addButton(self.letter_mode)
        mode_layout.addWidget(self.letter_mode)

        self.header_mode = QRadioButton("По заголовкам столбцов")
        self.header_mode.toggled.connect(self.on_mode_changed)
        self.mode_group.addButton(self.header_mode)
        mode_layout.addWidget(self.header_mode)

        layout.addWidget(mode_group)

        # Область сопоставления
        self.mapping_widget = QWidget()
        self.mapping_layout = QVBoxLayout(self.mapping_widget)
        layout.addWidget(self.mapping_widget)

        self.create_mapping_interface()

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выбери Excel", "", "Excel файлы (*.xlsx *.xls)"
        )
        if not file_path:
            return

        try:
            structure = get_excel_structure(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка чтения файла:\n{e}")
            return

        self.file_path = file_path
        self.file_structure = structure

        filename = os.path.basename(file_path)
        if len(filename) > 25:
            filename = filename[:22] + "..."
        self.file_btn.setText(filename)

        self.source_sheet_combo.clear()
        self.source_sheet_combo.addItems(list(structure.keys()))
        self.source_sheet_combo.setEnabled(True)

        self.update_title()
        self.create_mapping_interface()

    def update_title(self):
        if self.file_path:
            filename = os.path.basename(self.file_path)
            if len(filename) > 20:
                filename = filename[:17] + "..."

            source_sheet = self.source_sheet_combo.currentText()
            target_sheet = self.target_sheet_combo.currentText()

            title = f"{filename}"
            if source_sheet and target_sheet:
                title += f" [{source_sheet} → {target_sheet}]"

            self.title_label.setText(title)

    def on_mode_changed(self):
        self.create_mapping_interface()

    def create_mapping_interface(self):
        # Очищаем старый интерфейс
        for i in reversed(range(self.mapping_layout.count())):
            self.mapping_layout.itemAt(i).widget().setParent(None)

        if not self.file_path:
            info_label = QLabel("Сначала выберите файл")
            set_label_role(info_label, "muted")
            self.mapping_layout.addWidget(info_label)
            return

        source_sheet = self.source_sheet_combo.currentText()
        target_sheet = self.target_sheet_combo.currentText()

        if not source_sheet or not target_sheet:
            return

        source_headers = self.file_structure.get(source_sheet, [])
        target_headers = self.main_structure.get(target_sheet, [])

        if self.letter_mode.isChecked():
            self.create_letter_mapping(source_headers, target_headers)
        else:
            self.create_header_mapping(source_headers, target_headers)

    def create_letter_mapping(self, source_headers, target_headers):
        mapping_frame = QFrame()
        mapping_frame.setFrameStyle(QFrame.StyledPanel)
        mapping_layout = QVBoxLayout(mapping_frame)

        mapping_layout.addWidget(QLabel("Сопоставление столбцов (источник → назначение):"))

        # Создаем поля для ввода
        self.letter_mappings = []

        grid = QGridLayout()
        grid.addWidget(QLabel("Источник"), 0, 0)
        grid.addWidget(QLabel("→"), 0, 1)
        grid.addWidget(QLabel("Назначение"), 0, 2)
        grid.addWidget(QLabel(""), 0, 3)  # Для кнопки удаления

        # Добавляем первую строку
        self.add_letter_mapping_row(grid, 1)

        mapping_layout.addLayout(grid)

        # Кнопка добавления
        add_btn = QPushButton("+ Добавить сопоставление")
        add_btn.clicked.connect(lambda: self.add_letter_mapping_row(grid, len(self.letter_mappings) + 1))
        set_button_variant(add_btn, "secondary")
        mapping_layout.addWidget(add_btn)

        self.mapping_layout.addWidget(mapping_frame)

    def add_letter_mapping_row(self, grid, row):
        source_edit = QLineEdit()
        source_edit.setPlaceholderText("A")
        source_edit.setMaximumWidth(50)

        target_edit = QLineEdit()
        target_edit.setPlaceholderText("A")
        target_edit.setMaximumWidth(50)

        remove_btn = QPushButton("✕")
        remove_btn.setFixedSize(20, 20)
        set_button_variant(remove_btn, "danger")
        set_button_shape(remove_btn, "circle")
        remove_btn.clicked.connect(lambda: self.remove_letter_mapping_row(source_edit, target_edit, remove_btn))

        grid.addWidget(source_edit, row, 0)
        grid.addWidget(QLabel("→"), row, 1)
        grid.addWidget(target_edit, row, 2)
        grid.addWidget(remove_btn, row, 3)

        self.letter_mappings.append((source_edit, target_edit, remove_btn))

    def remove_letter_mapping_row(self, source_edit, target_edit, remove_btn):
        if len(self.letter_mappings) > 1:
            source_edit.setParent(None)
            target_edit.setParent(None)
            remove_btn.setParent(None)
            self.letter_mappings = [(s, t, r) for s, t, r in self.letter_mappings if s != source_edit]

    def create_header_mapping(self, source_headers, target_headers):
        mapping_frame = QFrame()
        mapping_frame.setFrameStyle(QFrame.StyledPanel)
        mapping_layout = QVBoxLayout(mapping_frame)

        mapping_layout.addWidget(QLabel("Сопоставление по заголовкам:"))

        # Создаем комбобоксы для сопоставления
        self.header_mappings = []

        grid = QGridLayout()
        grid.addWidget(QLabel("Источник"), 0, 0)
        grid.addWidget(QLabel("→"), 0, 1)
        grid.addWidget(QLabel("Назначение"), 0, 2)
        grid.addWidget(QLabel(""), 0, 3)  # Для кнопки удаления

        # Добавляем первую строку
        self.add_header_mapping_row(grid, 1, source_headers, target_headers)

        mapping_layout.addLayout(grid)

        # Кнопка добавления
        add_btn = QPushButton("+ Добавить сопоставление")
        add_btn.clicked.connect(
            lambda: self.add_header_mapping_row(grid, len(self.header_mappings) + 1, source_headers, target_headers))
        set_button_variant(add_btn, "secondary")
        mapping_layout.addWidget(add_btn)

        # Кнопка автосопоставления
        auto_btn = QPushButton("Автосопоставить")
        auto_btn.clicked.connect(lambda: self.auto_map_headers(source_headers, target_headers))
        set_button_variant(auto_btn, "secondary")
        mapping_layout.addWidget(auto_btn)

        self.mapping_layout.addWidget(mapping_frame)

    def add_header_mapping_row(self, grid, row, source_headers, target_headers):
        source_combo = QComboBox()
        source_combo.addItem("")  # Пустой элемент
        source_combo.addItems([f"{get_column_letter(i + 1) if get_column_letter else chr(65 + i)}: {h}"
                               for i, h in enumerate(source_headers)])

        target_combo = QComboBox()
        target_combo.addItem("")  # Пустой элемент
        target_combo.addItems([f"{get_column_letter(i + 1) if get_column_letter else chr(65 + i)}: {h}"
                               for i, h in enumerate(target_headers)])

        remove_btn = QPushButton("✕")
        remove_btn.setFixedSize(20, 20)
        set_button_variant(remove_btn, "danger")
        set_button_shape(remove_btn, "circle")
        remove_btn.clicked.connect(lambda: self.remove_header_mapping_row(source_combo, target_combo, remove_btn))

        grid.addWidget(source_combo, row, 0)
        grid.addWidget(QLabel("→"), row, 1)
        grid.addWidget(target_combo, row, 2)
        grid.addWidget(remove_btn, row, 3)

        self.header_mappings.append((source_combo, target_combo, remove_btn))

    def remove_header_mapping_row(self, source_combo, target_combo, remove_btn):
        if len(self.header_mappings) > 1:
            source_combo.setParent(None)
            target_combo.setParent(None)
            remove_btn.setParent(None)
            self.header_mappings = [(s, t, r) for s, t, r in self.header_mappings if s != source_combo]

    def auto_map_headers(self, source_headers, target_headers):
        # Очищаем текущие сопоставления
        for source_combo, target_combo, _ in self.header_mappings:
            source_combo.setCurrentIndex(0)
            target_combo.setCurrentIndex(0)

        # Находим совпадающие заголовки
        mapped_pairs = []
        for i, src_header in enumerate(source_headers):
            if src_header.strip():
                for j, tgt_header in enumerate(target_headers):
                    if src_header == tgt_header:
                        mapped_pairs.append((i, j))
                        break

        # Применяем сопоставления
        for idx, (src_idx, tgt_idx) in enumerate(mapped_pairs):
            if idx < len(self.header_mappings):
                self.header_mappings[idx][0].setCurrentIndex(src_idx + 1)  # +1 из-за пустого элемента
                self.header_mappings[idx][1].setCurrentIndex(tgt_idx + 1)

    def get_mapping(self):
        if not self.file_path:
            return None

        source_sheet = self.source_sheet_combo.currentText()
        target_sheet = self.target_sheet_combo.currentText()

        if not source_sheet or not target_sheet:
            return None

        source_columns = []
        target_columns = []

        if self.letter_mode.isChecked():
            # Обработка буквенных обозначений
            for source_edit, target_edit, _ in self.letter_mappings:
                src_text = source_edit.text().strip().upper()
                tgt_text = target_edit.text().strip().upper()

                if src_text and tgt_text:
                    # Просто используем буквенные обозначения как есть
                    source_columns.append(src_text)
                    target_columns.append(tgt_text)
        else:
            # Обработка заголовков
            for source_combo, target_combo, _ in self.header_mappings:
                src_idx = source_combo.currentIndex()
                tgt_idx = target_combo.currentIndex()

                if src_idx > 0 and tgt_idx > 0:  # Не пустые значения
                    source_headers = self.file_structure.get(source_sheet, [])
                    target_headers = self.main_structure.get(target_sheet, [])

                    if src_idx - 1 < len(source_headers) and tgt_idx - 1 < len(target_headers):
                        source_columns.append(source_headers[src_idx - 1])
                        target_columns.append(target_headers[tgt_idx - 1])

        if not source_columns or not target_columns:
            return None

        return {
            "source": self.file_path,
            "source_columns": source_columns,
            "target_sheet": target_sheet,
            "target_columns": target_columns
        }


class MergeMappingDialog(QDialog):
    def __init__(self, main_excel_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки объединения")
        self.resize(750, 500)
        self.setMinimumSize(600, 400)

        try:
            self.main_structure = get_excel_structure(main_excel_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка чтения главного файла:\n{e}")
            self.main_structure = {}

        self.cards = []
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Заголовок
        title = QLabel("Настройка объединения Excel файлов")
        title.setAlignment(Qt.AlignCenter)
        set_label_role(title, "heading")
        layout.addWidget(title)

        # Скролл область для карточек
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        self.scroll_content = QWidget()
        self.cards_layout = QVBoxLayout(self.scroll_content)
        self.cards_layout.setSpacing(8)
        self.cards_layout.setContentsMargins(5, 5, 5, 5)
        scroll.setWidget(self.scroll_content)
        layout.addWidget(scroll)

        # Кнопки управления
        controls_layout = QHBoxLayout()

        self.add_btn = QPushButton("+ Добавить сопоставление")
        self.add_btn.clicked.connect(self.add_card)
        set_button_variant(self.add_btn, "secondary")
        controls_layout.addWidget(self.add_btn)

        self.apply_first_btn = QPushButton("Применить настройки первого ко всем")
        self.apply_first_btn.clicked.connect(self.apply_first_to_all)
        set_button_variant(self.apply_first_btn, "secondary")
        controls_layout.addWidget(self.apply_first_btn)

        controls_layout.addStretch()
        layout.addLayout(controls_layout)

        # Кнопки диалога
        dialog_btns = QHBoxLayout()
        dialog_btns.addStretch()

        self.ok_btn = QPushButton("Готово")
        self.ok_btn.clicked.connect(self.accept)
        self.ok_btn.setMinimumWidth(100)
        set_button_variant(self.ok_btn, "primary")

        self.cancel_btn = QPushButton("Отмена")
        self.cancel_btn.clicked.connect(self.reject)
        self.cancel_btn.setMinimumWidth(100)
        set_button_variant(self.cancel_btn, "secondary")

        dialog_btns.addWidget(self.ok_btn)
        dialog_btns.addWidget(self.cancel_btn)
        layout.addLayout(dialog_btns)

    def add_card(self):
        card = MappingCard(self.main_structure, self)
        card.remove_btn.clicked.connect(lambda: self.remove_card(card))
        self.cards.append(card)
        self.cards_layout.addWidget(card)

    def add_row_with_file(self, file_path):
        card = MappingCard(self.main_structure, self)
        card.remove_btn.clicked.connect(lambda: self.remove_card(card))

        try:
            structure = get_excel_structure(file_path)
            card.file_path = file_path
            card.file_structure = structure

            filename = os.path.basename(file_path)
            if len(filename) > 25:
                filename = filename[:22] + "..."
            card.file_btn.setText(filename)

            card.source_sheet_combo.clear()
            card.source_sheet_combo.addItems(list(structure.keys()))
            card.source_sheet_combo.setEnabled(True)

            card.update_title()
            card.create_mapping_interface()
        except Exception as e:
            QMessageBox.warning(self, "Предупреждение", f"Не удалось загрузить {file_path}: {e}")

        self.cards.append(card)
        self.cards_layout.addWidget(card)

    def remove_card(self, card):
        self.cards.remove(card)
        card.setParent(None)
        card.deleteLater()

    def apply_first_to_all(self):
        if not self.cards:
            QMessageBox.warning(self, "Предупреждение", "Нет карточек для применения настроек.")
            return

        first_card = self.cards[0]

        if not first_card.file_path:
            QMessageBox.warning(self, "Предупреждение", "В первой карточке не выбран файл.")
            return

        # Получаем настройки из первой карточки
        first_settings = self.get_first_card_settings(first_card)
        if not first_settings:
            QMessageBox.warning(self, "Предупреждение", "В первой карточке нет настроенных сопоставлений.")
            return

        # Применяем к остальным карточкам
        applied_count = 0
        for card in self.cards[1:]:
            if card.file_path:
                self.apply_settings_to_card(card, first_settings)
                applied_count += 1

        if applied_count > 0:
            QMessageBox.information(self, "Успех", f"Настройки применены к {applied_count} карточкам.")
        else:
            QMessageBox.warning(self, "Предупреждение", "Нет карточек с выбранными файлами для применения.")

    def get_first_card_settings(self, card):
        settings = {
            'mode': 'letter' if card.letter_mode.isChecked() else 'header',
            'target_sheet': card.target_sheet_combo.currentText(),
            'mappings': []
        }

        if card.letter_mode.isChecked():
            for source_edit, target_edit, _ in card.letter_mappings:
                src_text = source_edit.text().strip()
                tgt_text = target_edit.text().strip()
                if src_text and tgt_text:
                    settings['mappings'].append((src_text, tgt_text))
        else:
            for source_combo, target_combo, _ in card.header_mappings:
                src_idx = source_combo.currentIndex()
                tgt_idx = target_combo.currentIndex()
                if src_idx > 0 and tgt_idx > 0:
                    src_text = source_combo.currentText()
                    tgt_text = target_combo.currentText()
                    settings['mappings'].append((src_text, tgt_text))

        return settings if settings['mappings'] else None

    def apply_settings_to_card(self, card, settings):
        # Устанавливаем режим
        if settings['mode'] == 'letter':
            card.letter_mode.setChecked(True)
        else:
            card.header_mode.setChecked(True)

        # Устанавливаем целевой лист
        target_idx = card.target_sheet_combo.findText(settings['target_sheet'])
        if target_idx >= 0:
            card.target_sheet_combo.setCurrentIndex(target_idx)

        # Пересоздаем интерфейс сопоставления
        card.create_mapping_interface()

        # Применяем сопоставления
        if settings['mode'] == 'letter':
            # Очищаем текущие сопоставления
            for source_edit, target_edit, remove_btn in card.letter_mappings[1:]:
                source_edit.setParent(None)
                target_edit.setParent(None)
                remove_btn.setParent(None)
            card.letter_mappings = card.letter_mappings[:1]

            # Применяем новые
            for i, (src_text, tgt_text) in enumerate(settings['mappings']):
                if i == 0 and card.letter_mappings:
                    # Используем первую существующую строку
                    card.letter_mappings[0][0].setText(src_text)
                    card.letter_mappings[0][1].setText(tgt_text)
                else:
                    # Добавляем новые строки
                    grid = card.mapping_widget.layout().itemAt(0).widget().layout().itemAt(1).layout()
                    card.add_letter_mapping_row(grid, len(card.letter_mappings) + 1)
                    card.letter_mappings[-1][0].setText(src_text)
                    card.letter_mappings[-1][1].setText(tgt_text)
        else:
            # Для режима заголовков
            source_headers = card.file_structure.get(card.source_sheet_combo.currentText(), [])
            target_headers = card.main_structure.get(card.target_sheet_combo.currentText(), [])

            # Очищаем текущие сопоставления
            for source_combo, target_combo, remove_btn in card.header_mappings[1:]:
                source_combo.setParent(None)
                target_combo.setParent(None)
                remove_btn.setParent(None)
            card.header_mappings = card.header_mappings[:1]

            # Применяем новые
            for i, (src_text, tgt_text) in enumerate(settings['mappings']):
                if i == 0 and card.header_mappings:
                    # Используем первую существующую строку
                    src_idx = card.header_mappings[0][0].findText(src_text)
                    tgt_idx = card.header_mappings[0][1].findText(tgt_text)
                    if src_idx >= 0:
                        card.header_mappings[0][0].setCurrentIndex(src_idx)
                    if tgt_idx >= 0:
                        card.header_mappings[0][1].setCurrentIndex(tgt_idx)
                else:
                    # Добавляем новые строки
                    grid = card.mapping_widget.layout().itemAt(0).widget().layout().itemAt(2).layout()
                    card.add_header_mapping_row(grid, len(card.header_mappings) + 1, source_headers, target_headers)
                    src_idx = card.header_mappings[-1][0].findText(src_text)
                    tgt_idx = card.header_mappings[-1][1].findText(tgt_text)
                    if src_idx >= 0:
                        card.header_mappings[-1][0].setCurrentIndex(src_idx)
                    if tgt_idx >= 0:
                        card.header_mappings[-1][1].setCurrentIndex(tgt_idx)

    def get_mappings(self):
        result = []
        for card in self.cards:
            mapping = card.get_mapping()
            if mapping:
                result.append(mapping)
        return result
