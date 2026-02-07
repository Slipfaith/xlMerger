# -*- coding: utf-8 -*-
from PySide6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QLabel, QPushButton, QMessageBox,
    QApplication, QProgressDialog
)
from PySide6.QtCore import Qt
from utils.i18n import tr, i18n
from core.drag_drop import DragDropLineEdit
from core.split_excel import split_excel_multiple_sheets
from gui.split_mapping_dialog import SplitMappingDialog
from .style_system import set_button_variant
from openpyxl import load_workbook


class SplitTab(QWidget):
    def __init__(self):
        super().__init__()
        # path to selected Excel file
        self.excel_path = ''
        # mapping per sheet: {sheet: (src, [targets], [extras])}
        self.sheet_mappings: dict[str, tuple[str, list[str], list[str]]] = {}
        self.init_ui()
        i18n.language_changed.connect(self.retranslate_ui)
        self.retranslate_ui()

    def init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(30)

        left = QVBoxLayout()
        left.setSpacing(15)
        self.file_input = DragDropLineEdit(mode='file')
        self.file_input.setPlaceholderText(tr("Перетащи сюда эксель"))
        self.file_input.fileSelected.connect(self.on_file_selected)
        self.file_input.setMinimumHeight(40)
        left.addWidget(self.file_input)

        self.config_btn = QPushButton(tr("Настроить"))
        set_button_variant(self.config_btn, "secondary")
        self.config_btn.clicked.connect(self.open_mapping_dialog)

        self.split_btn = QPushButton(tr("Разделить"))
        set_button_variant(self.split_btn, "orange")
        self.split_btn.clicked.connect(self.run_split)

        action_row = QHBoxLayout()
        action_row.setSpacing(10)
        action_row.addWidget(self.config_btn)
        action_row.addWidget(self.split_btn)

        left.addStretch()
        left.addLayout(action_row)

        layout.addLayout(left)

        right = QVBoxLayout()
        self.current_label = QLabel(tr("Текущая настройка: —"))
        self.current_label.setWordWrap(True)
        self.current_label.hide()
        right.addStretch()

        layout.addLayout(right)

        # Use application defaults for styling to match other tabs

    def on_file_selected(self, path):
        """Store selected Excel path and reset mappings."""
        self.excel_path = path
        # reset current selection
        self.sheet_mappings = {}
        self.current_label.setText(tr("Текущая настройка: —"))

    def open_mapping_dialog(self):
        if not self.excel_path:
            QMessageBox.critical(self, tr("Ошибка"), tr("Выберите файл Excel."))
            return
        wb = load_workbook(self.excel_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        dialog = SplitMappingDialog(self.excel_path, sheets, self)
        if dialog.exec():
            self.sheet_mappings = dialog.get_selection()
            self._update_current_label()

    def run_split(self):
        if not self.excel_path:
            QMessageBox.critical(self, tr("Ошибка"), tr("Выберите файл Excel."))
            return
        if not self.sheet_mappings:
            QMessageBox.critical(self, tr("Ошибка"), tr("Сначала настройте листы."))
            return
        try:
            progress = QProgressDialog(tr("Сохранение..."), tr("Отмена"), 0, 0, self)
            progress.setWindowTitle(tr("Прогресс"))
            progress.setWindowModality(Qt.ApplicationModal)
            progress.setMinimumDuration(0)
            progress.show()
            QApplication.processEvents()

            def cb(i, total, name):
                progress.setMaximum(total)
                progress.setValue(i)
                progress.setLabelText(name)
                QApplication.processEvents()

            cfg = {
                sheet: (
                    src,
                    targets if targets else None,
                    extras,
                )
                for sheet, (src, targets, extras) in self.sheet_mappings.items()
            }

            split_excel_multiple_sheets(
                self.excel_path,
                cfg,
                progress_callback=cb,
            )

            progress.close()
            QMessageBox.information(self, tr("Успех"), tr("Файлы успешно сохранены."))
        except Exception as e:
            QMessageBox.critical(self, tr("Ошибка"), str(e))

    def retranslate_ui(self):
        self.setWindowTitle(tr("xlSpliter"))
        self.split_btn.setText(tr("Разделить"))
        self.config_btn.setText(tr("Настроить"))
        self.file_input.setPlaceholderText(tr("Перетащи сюда эксель"))
        self._update_current_label()
        # update labels - they are static but to refresh we need to re-add them? Not necessary

    def _update_current_label(self):
        if not self.sheet_mappings:
            self.current_label.setText(tr("Текущая настройка: —"))
            return

        parts = []
        for sheet, (src, targets, extras) in self.sheet_mappings.items():
            tgts = ', '.join(targets) if targets else '—'
            ex = ', '.join(extras) if extras else '—'
            parts.append(
                f"<b>{sheet}</b>: {tr('Источник')}: {src}; {tr('Цели')}: {tgts}; {tr('Доп')}: {ex}"
            )
        html = '<br>'.join(parts)
        self.current_label.setText(tr('Текущая настройка:') + '<br>' + html)

