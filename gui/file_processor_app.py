import os
import sys
import json
import traceback
import hashlib
import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QFileDialog, QVBoxLayout, QHBoxLayout,
    QTableView, QHeaderView, QComboBox, QMessageBox, QProgressBar, QScrollArea, QFrame, QCheckBox, QListWidget,
    QListWidgetItem, QRadioButton, QGridLayout, QDialog, QListView, QStackedWidget
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QStandardItemModel, QStandardItem
from openpyxl import load_workbook
import openpyxl.utils as utils
from openpyxl.styles import PatternFill

from gui.excel_previewer import ExcelPreviewer
from gui.excel_file_selector import ExcelFileSelector
from utils import excel_column_to_index

class FileProcessorApp(QWidget):
    copyingStarted = Signal()

    def __init__(self):
        super().__init__()

        self.folder_path = ''
        self.excel_file_path = ''
        self.copy_column = ''
        self.selected_files = None
        self.selected_sheets = []
        self.columns = {}
        self.header_row = {}
        self.sheet_to_column = {}
        self.folder_to_column = {}
        self.file_to_column = {}
        self.skip_first_row = False
        self.copy_by_row_number = False
        self.workbook = None

        self.stack = QStackedWidget(self)
        self.page_main = self.create_main_page()
        self.stack.addWidget(self.page_main)

        layout = QVBoxLayout()
        layout.addWidget(self.stack)
        self.setLayout(layout)

        self.setWindowTitle(self.tr("Обработка файлов"))
        self.center_window()
        self.setGeometry(300, 300, 600, 400)
        self.show()

    # ======= Главная страница =======
    def create_main_page(self):
        widget = QWidget()
        layout = QVBoxLayout()

        layout.addLayout(self.create_folder_selection_layout())
        layout.addLayout(self.create_excel_selection_layout())
        layout.addLayout(self.create_sheet_selection_layout())
        layout.addLayout(self.create_copy_column_layout())
        layout.addWidget(self.create_skip_first_row_checkbox())
        layout.addLayout(self.create_copy_method_selection_layout())
        layout.addWidget(self.create_preview_button(), alignment=Qt.AlignRight)
        layout.addWidget(self.create_process_button(), alignment=Qt.AlignCenter)

        widget.setLayout(layout)
        return widget

    def create_folder_selection_layout(self):
        layout = QHBoxLayout()
        self.folder_entry = self.create_folder_entry()
        folder_button = QPushButton(self.tr("Обзор"), self)
        folder_button.clicked.connect(self.select_folder)

        layout.addWidget(QLabel(self.tr("Путь к папке с переводами:")))
        layout.addWidget(self.folder_entry)
        layout.addWidget(folder_button)
        return layout

    def create_folder_entry(self):
        folder_entry = QLineEdit(self)
        folder_entry.setPlaceholderText(self.tr("Перетащите сюда папку"))
        folder_entry.setAcceptDrops(True)
        folder_entry.dragEnterEvent = self.dragEnterEvent
        folder_entry.dropEvent = self.dropEvent
        return folder_entry

    def create_excel_selection_layout(self):
        layout = QHBoxLayout()
        self.excel_file_entry = self.create_excel_file_entry()
        excel_button = QPushButton(self.tr("Обзор"), self)
        excel_button.clicked.connect(self.select_excel_file)

        layout.addWidget(QLabel(self.tr("Путь к файлу Excel:")))
        layout.addWidget(self.excel_file_entry)
        layout.addWidget(excel_button)
        return layout

    def create_excel_file_entry(self):
        excel_file_entry = QLineEdit(self)
        excel_file_entry.setPlaceholderText(self.tr("Перетащите сюда xlsx или нажмите \"Обзор\""))
        excel_file_entry.setAcceptDrops(True)
        excel_file_entry.dragEnterEvent = self.dragEnterEventExcel
        excel_file_entry.dropEvent = self.dropEventExcel
        return excel_file_entry

    def create_sheet_selection_layout(self):
        layout = QVBoxLayout()
        self.sheet_list = QListWidget(self)
        self.sheet_list.setSelectionMode(QListWidget.MultiSelection)
        self.sheet_list.setFixedHeight(100)

        button_layout = QHBoxLayout()
        deselect_all_button = QPushButton(self.tr("Снять выделение со всех"), self)
        select_all_button = QPushButton(self.tr("Выбрать все"), self)
        deselect_all_button.clicked.connect(self.deselect_all_sheets)
        select_all_button.clicked.connect(self.select_all_sheets)
        button_layout.addWidget(deselect_all_button)
        button_layout.addWidget(select_all_button)

        layout.addWidget(QLabel(self.tr("Выберите листы:")))
        layout.addWidget(self.sheet_list)
        layout.addLayout(button_layout)
        return layout

    def create_copy_column_layout(self):
        layout = QHBoxLayout()
        self.copy_column_entry = QLineEdit(self)
        self.copy_column_entry.setMaximumWidth(100)
        copy_column_label = QLabel(self.tr("Столбец для копирования:"))
        copy_column_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        layout.addWidget(copy_column_label)
        layout.addWidget(self.copy_column_entry)
        return layout

    def create_skip_first_row_checkbox(self):
        self.skip_first_row_checkbox = QCheckBox(self.tr("Первая строка - заголовок в переводах"), self)
        self.skip_first_row_checkbox.stateChanged.connect(self.toggle_skip_first_row)
        return self.skip_first_row_checkbox

    def create_copy_method_selection_layout(self):
        layout = QHBoxLayout()
        self.copy_by_matching_radio = QRadioButton(self.tr("Нет пустых/скрытых строк в xlsx"), self)
        self.copy_by_matching_radio.setChecked(True)
        self.copy_by_matching_radio.toggled.connect(self.toggle_copy_method)
        self.copy_by_row_number_radio = QRadioButton(self.tr("Есть пустые/скрытые строки в xlsx"), self)
        layout.addWidget(self.copy_by_matching_radio)
        layout.addWidget(self.copy_by_row_number_radio)
        return layout

    def create_preview_button(self):
        preview_button = QPushButton(self.tr("Превью"), self)
        preview_button.clicked.connect(self.select_excel_file_for_preview)
        return preview_button

    def create_process_button(self):
        process_button = QPushButton(self.tr("Начать"), self)
        process_button.setStyleSheet("background-color: #f47929; color: white;")
        process_button.clicked.connect(self.process_files)
        return process_button

    # Drag&Drop
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls and urls[0].isLocalFile():
            file_paths = self.get_dropped_files(urls)
            if file_paths:
                self.folder_path = os.path.dirname(file_paths[0])
                self.folder_entry.setText(self.folder_path)
                self.selected_files = [os.path.basename(fp) for fp in file_paths]
            elif os.path.isdir(urls[0].toLocalFile()):
                self.folder_path = urls[0].toLocalFile()
                self.folder_entry.setText(self.folder_path)
                self.selected_files = None
            else:
                QMessageBox.warning(self, self.tr("Предупреждение"), self.tr("Неверный тип файла. Поддерживаются только файлы Excel."))

    def get_dropped_files(self, urls):
        return [url.toLocalFile() for url in urls if url.toLocalFile().endswith(('.xlsx', '.xls'))]

    def dragEnterEventExcel(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEventExcel(self, event):
        urls = event.mimeData().urls()
        if urls and urls[0].isLocalFile():
            file_path = urls[0].toLocalFile()
            if file_path.endswith(('.xlsx', '.xls')):
                self.excel_file_entry.setText(file_path)
                self.excel_file_path = file_path
                self.load_sheet_names()

    def select_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, self.tr("Выберите папку"))
        if folder_path:
            self.folder_entry.setText(folder_path)
            self.folder_path = folder_path

    def select_excel_file(self):
        excel_file_path, _ = QFileDialog.getOpenFileName(self, self.tr("Выберите файл Excel"), '', self.tr("Excel файлы (*.xlsx *.xls)"))
        if excel_file_path:
            self.excel_file_entry.setText(excel_file_path)
            self.excel_file_path = excel_file_path
            self.load_sheet_names()

    def load_sheet_names(self):
        try:
            workbook = load_workbook(self.excel_file_path, read_only=True)
            self.sheet_names = workbook.sheetnames
            self.populate_sheet_list()
        except Exception as e:
            self.log_error(e)
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Не удалось загрузить листы из файла Excel: ") + str(e))

    def populate_sheet_list(self):
        self.sheet_list.clear()
        for sheet_name in self.sheet_names:
            item = QListWidgetItem(sheet_name)
            item.setCheckState(Qt.Unchecked)
            self.sheet_list.addItem(item)

    def deselect_all_sheets(self):
        self.set_sheet_selection_state(Qt.Unchecked)

    def select_all_sheets(self):
        self.set_sheet_selection_state(Qt.Checked)

    def set_sheet_selection_state(self, state):
        for index in range(self.sheet_list.count()):
            item = self.sheet_list.item(index)
            item.setCheckState(state)

    def process_files(self):
        self.folder_path = self.folder_entry.text()
        self.excel_file_path = self.excel_file_entry.text()
        self.copy_column = self.copy_column_entry.text()

        if not self.validate_paths_and_column():
            return

        self.selected_sheets = [self.sheet_list.item(index).text()
                                for index in range(self.sheet_list.count())
                                if self.sheet_list.item(index).checkState() == Qt.Checked]

        if not self.selected_sheets:
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Выберите хотя бы один лист."))
            return

        try:
            self.workbook = load_workbook(self.excel_file_path)
            self.columns = {}
            self.header_row = {}
            self.go_to_header_page()
        except Exception as e:
            self.log_error(e)
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Произошла ошибка при обработке файлов: ") + str(e))

    def validate_paths_and_column(self):
        if not os.path.exists(self.folder_path):
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Указанная папка не существует."))
            return False

        if not os.path.exists(self.excel_file_path):
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Указанный файл Excel не существует."))
            return False

        if not self.copy_column:
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Укажите столбец для копирования."))
            return False

        return True

    def log_error(self, error):
        log_file_path = os.path.join(os.path.dirname(__file__), 'error_log.txt')
        with open(log_file_path, 'a', encoding='utf-8') as log_file:
            log_file.write(self.tr("Ошибка: ") + str(error) + "\n")
            log_file.write(traceback.format_exc())
            log_file.write("\n")

    def center_window(self, window=None):
        if window is None:
            window = self
        screen = QApplication.primaryScreen()
        if screen:
            rect = screen.availableGeometry()
            x = (rect.width() - self.width()) // 2
            y = (rect.height() - self.height()) // 2
            window.move(x, y)

    def toggle_skip_first_row(self, state):
        self.skip_first_row = state == Qt.Checked

    def toggle_copy_method(self):
        self.copy_by_row_number = self.copy_by_row_number_radio.isChecked()
        self.skip_first_row_checkbox.setEnabled(not self.copy_by_row_number)
        if self.copy_by_row_number:
            self.skip_first_row_checkbox.setChecked(False)

    def select_excel_file_for_preview(self):
        if not self.folder_path:
            QMessageBox.warning(self, self.tr("Предупреждение"), self.tr("Сначала выберите папку с переводами."))
            return

        dialog = ExcelFileSelector(self.folder_path, self.selected_files)
        if dialog.exec():
            selected_file = dialog.selected_file
            if selected_file:
                self.preview_window = ExcelPreviewer(selected_file)
                self.preview_window.show()

    # ======= Страница выбора строки заголовка =======
    def create_header_page(self):
        page = QWidget()
        page.setWindowTitle(self.tr("Выберите строку заголовка"))
        layout = QVBoxLayout()
        layout.addWidget(QLabel(self.tr("Выберите номер строки заголовка в файле Excel для каждого листа:")))

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        scroll_layout = QGridLayout(scroll_content)
        scroll_content.setLayout(scroll_layout)
        scroll_area.setWidget(scroll_content)

        self.sheet_to_header_row = {}
        for row, sheet_name in enumerate(self.selected_sheets):
            sheet_label = QLabel(sheet_name)
            header_row_combobox = QComboBox()
            header_row_combobox.setMaximumWidth(100)
            header_row_combobox.addItems([str(i) for i in range(1, 11)])
            header_row_combobox.setCurrentIndex(0)
            scroll_layout.addWidget(sheet_label, row, 0)
            scroll_layout.addWidget(header_row_combobox, row, 1)
            self.sheet_to_header_row[sheet_name] = header_row_combobox

        layout.addWidget(scroll_area)

        preview_button = QPushButton(self.tr("Превью"), page)
        preview_button.clicked.connect(self.preview_target_excel)
        layout.addWidget(preview_button, alignment=Qt.AlignCenter)

        button_layout = QHBoxLayout()
        back_button = QPushButton(self.tr("Назад"))
        next_button = QPushButton(self.tr("Далее"))
        back_button.clicked.connect(self.go_to_main_page)
        next_button.clicked.connect(self.load_columns_and_go_to_sheet_column)
        button_layout.addWidget(back_button)
        button_layout.addWidget(next_button)
        layout.addLayout(button_layout)

        page.setLayout(layout)
        return page

    def go_to_header_page(self):
        self.page_header = self.create_header_page()
        self.stack.addWidget(self.page_header)
        self.stack.setCurrentWidget(self.page_header)

    def preview_target_excel(self):
        if not self.excel_file_path:
            QMessageBox.warning(self, self.tr("Предупреждение"), self.tr("Сначала выберите файл Excel."))
            return
        self.preview_window = ExcelPreviewer(self.excel_file_path)
        self.preview_window.show()

    def go_to_main_page(self):
        self.stack.setCurrentWidget(self.page_main)

    def load_columns_and_go_to_sheet_column(self):
        try:
            for sheet_name, combobox in self.sheet_to_header_row.items():
                header_row_index = int(combobox.currentText()) - 1
                self.header_row[sheet_name] = header_row_index
                sheet = self.workbook[sheet_name]
                self.columns[sheet_name] = [cell.value for cell in sheet[header_row_index + 1]]
            self.go_to_sheet_column_page()
        except Exception as e:
            self.log_error(e)
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Произошла ошибка при загрузке столбцов: ") + str(e))

    # ======= Страница соответствия лист-столбец =======
    def create_sheet_column_page(self):
        page = QWidget()
        page.setWindowTitle(self.tr("Соответствие лист-столбец"))
        layout = QVBoxLayout()
        layout.addWidget(QLabel(self.tr("Из какого столбца на каждом листе должны копироваться переводы?")))

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        scroll_layout = QGridLayout(scroll_content)
        scroll_content.setLayout(scroll_layout)
        scroll_area.setWidget(scroll_content)

        self.sheet_to_column = {}
        default_column = self.copy_column
        for row, sheet_name in enumerate(self.selected_sheets):
            sheet_label = QLabel(sheet_name)
            column_entry = QLineEdit()
            column_entry.setMaximumWidth(100)
            column_entry.setText(default_column)
            scroll_layout.addWidget(sheet_label, row, 0)
            scroll_layout.addWidget(column_entry, row, 1)
            self.sheet_to_column[sheet_name] = column_entry

        layout.addWidget(scroll_area)

        preview_button = QPushButton(self.tr("Превью"), page)
        preview_button.clicked.connect(self.select_excel_file_for_preview)
        layout.addWidget(preview_button, alignment=Qt.AlignCenter)

        button_layout = QHBoxLayout()
        back_button = QPushButton(self.tr("Назад"))
        next_button = QPushButton(self.tr("Далее"))
        back_button.clicked.connect(self.go_to_header_page)
        next_button.clicked.connect(self.go_to_match_page)
        button_layout.addWidget(back_button)
        button_layout.addWidget(next_button)
        layout.addLayout(button_layout)

        page.setLayout(layout)
        return page

    def go_to_sheet_column_page(self):
        self.page_sheet_column = self.create_sheet_column_page()
        self.stack.addWidget(self.page_sheet_column)
        self.stack.setCurrentWidget(self.page_sheet_column)

    # ======= Страница сопоставления (файл/папка) =======
    def create_match_page(self):
        page = QWidget()
        if self.are_all_items_files(os.listdir(self.folder_path)):
            page.setWindowTitle(self.tr("Соответствие файл-столбец"))
        else:
            page.setWindowTitle(self.tr("Соответствие папка-столбец"))
        layout = QVBoxLayout()
        if self.are_all_items_files(os.listdir(self.folder_path)):
            layout.addWidget(QLabel(self.tr("Сопоставьте имена файлов с именами столбцов:")))
        else:
            layout.addWidget(QLabel(self.tr("Сопоставьте имена папок с именами столбцов:")))

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        scroll_layout = QGridLayout(scroll_content)
        scroll_content.setLayout(scroll_layout)
        scroll_area.setWidget(scroll_content)

        if self.are_all_items_files(os.listdir(self.folder_path)):
            self.file_to_column = {}
            available_columns = [''] + list(set(sum([self.columns[sheet] for sheet in self.selected_sheets], [])))
            row = 0
            col = 0
            files_to_process = self.selected_files if self.selected_files else [
                os.path.join(self.folder_path, file_name) for file_name in os.listdir(self.folder_path)
                if file_name.endswith(('.xlsx', '.xls'))
            ]
            for file_path in files_to_process:
                file_name = os.path.basename(file_path)
                if self.selected_files and file_name not in self.selected_files:
                    continue
                if row >= 5:
                    row = 0
                    col += 2
                file_label = QLabel(file_name)
                file_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                column_combobox = QComboBox()
                column_combobox.setMaximumWidth(100)
                column_combobox.addItems(available_columns)
                scroll_layout.addWidget(file_label, row, col)
                scroll_layout.addWidget(column_combobox, row, col+1)
                self.file_to_column[file_name] = column_combobox
                row += 1
        else:
            self.folder_to_column = {}
            available_columns = [''] + list(set(sum([self.columns[sheet] for sheet in self.selected_sheets], [])))
            row = 0
            col = 0
            for folder_name in os.listdir(self.folder_path):
                lang_folder_path = os.path.join(self.folder_path, folder_name)
                if os.path.isdir(lang_folder_path):
                    if row >= 5:
                        row = 0
                        col += 2
                    folder_label = QLabel(folder_name)
                    folder_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    column_combobox = QComboBox()
                    column_combobox.setMaximumWidth(100)
                    column_combobox.addItems(available_columns)
                    scroll_layout.addWidget(folder_label, row, col)
                    scroll_layout.addWidget(column_combobox, row, col+1)
                    self.folder_to_column[folder_name] = column_combobox
                    row += 1

        layout.addWidget(scroll_area)

        button_layout = QHBoxLayout()
        save_button = QPushButton(self.tr("Сохранить настройки"))
        load_button = QPushButton(self.tr("Загрузить настройки"))
        back_button = QPushButton(self.tr("Назад"))
        next_button = QPushButton(self.tr("Далее"))
        save_button.clicked.connect(self.save_mapping_settings)
        load_button.clicked.connect(self.load_mapping_settings)
        back_button.clicked.connect(self.go_to_sheet_column_page)
        next_button.clicked.connect(self.go_to_confirmation_page)
        button_layout.addWidget(save_button)
        button_layout.addWidget(load_button)
        button_layout.addWidget(back_button)
        button_layout.addWidget(next_button)
        layout.addLayout(button_layout)

        page.setLayout(layout)
        return page

    def are_all_items_files(self, folder_content):
        return all(os.path.isfile(os.path.join(self.folder_path, item)) for item in folder_content)

    def go_to_match_page(self):
        self.page_match = self.create_match_page()
        self.stack.addWidget(self.page_match)
        self.stack.setCurrentWidget(self.page_match)

    # ======= Страница подтверждения сопоставления =======
    def create_confirmation_page(self):
        page = QWidget()
        page.setWindowTitle(self.tr("Подтверждение сопоставления"))
        layout = QVBoxLayout()
        layout.addWidget(QLabel(self.tr("Подтвердите сопоставление:")))

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        scroll_layout = QGridLayout(scroll_content)
        scroll_layout.setHorizontalSpacing(10)
        scroll_content.setLayout(scroll_layout)
        scroll_area.setWidget(scroll_content)

        items = self.get_sorted_items()
        row = 0
        col = 0
        for index, (name, combobox) in enumerate(items):
            scroll_layout.addWidget(QLabel(f"{name}: {combobox.currentText()}"), row, col)
            col += 1
            if col > 1:
                col = 0
                row += 1

        layout.addWidget(scroll_area)

        button_layout = QHBoxLayout()
        back_button = QPushButton(self.tr("Назад"))
        start_button = QPushButton(self.tr("Начать"))
        start_button.setStyleSheet("background-color: #f47929; color: white;")
        back_button.clicked.connect(self.go_to_match_page)
        start_button.clicked.connect(self.start_copying)
        button_layout.addWidget(back_button)
        button_layout.addWidget(start_button)
        layout.addLayout(button_layout)

        page.setLayout(layout)
        return page

    def go_to_confirmation_page(self):
        self.page_confirmation = self.create_confirmation_page()
        self.stack.addWidget(self.page_confirmation)
        self.stack.setCurrentWidget(self.page_confirmation)

    def get_sorted_items(self):
        if hasattr(self, 'file_to_column') and self.file_to_column:
            return sorted(self.file_to_column.items(), key=lambda x: (x[1].currentText() == "", x[0]))
        else:
            return sorted(self.folder_to_column.items(), key=lambda x: (x[1].currentText() == "", x[0]))

    # ======= Страница прогресса копирования =======
    def create_progress_page(self):
        page = QWidget()
        page.setWindowTitle(self.tr("Копирование переводов"))
        layout = QVBoxLayout()
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #f47929;
            }
        """)
        layout.addWidget(self.progress_bar)
        page.setLayout(layout)
        return page

    def go_to_progress_page(self):
        self.page_progress = self.create_progress_page()
        self.stack.addWidget(self.page_progress)
        self.stack.setCurrentWidget(self.page_progress)

    # ======= Страница завершения копирования =======
    def create_completion_page(self):
        page = QWidget()
        page.setWindowTitle(self.tr("Копирование завершено"))
        layout = QVBoxLayout()
        layout.addWidget(QLabel(self.tr("Файлы успешно сохранены.")))

        button_layout = QHBoxLayout()
        close_button = QPushButton(self.tr("Закрыть приложение"))
        restart_button = QPushButton(self.tr("Вернуться на главный экран"))
        close_button.clicked.connect(QApplication.instance().quit)
        restart_button.clicked.connect(self.return_to_main_screen)
        button_layout.addWidget(close_button)
        button_layout.addWidget(restart_button)
        layout.addLayout(button_layout)

        page.setLayout(layout)
        return page

    def go_to_completion_page(self):
        self.page_completion = self.create_completion_page()
        self.stack.addWidget(self.page_completion)
        self.stack.setCurrentWidget(self.page_completion)

    def return_to_main_screen(self):
        self.reset_to_initial_state()
        self.stack.setCurrentWidget(self.page_main)

    def reset_to_initial_state(self):
        if self.workbook:
            self.workbook.close()
            self.workbook = None
        self.folder_entry.clear()
        self.excel_file_entry.clear()
        self.copy_column_entry.clear()
        self.skip_first_row_checkbox.setChecked(False)
        self.copy_by_matching_radio.setChecked(True)
        self.sheet_list.clear()
        self.columns = {}
        self.header_row = {}
        self.sheet_to_column = {}
        self.folder_to_column = {}
        self.file_to_column = {}

    # ======= Сохранение/загрузка настроек сопоставления =======
    def save_mapping_settings(self):
        try:
            mapping = self.get_current_mapping()
            settings_path, _ = QFileDialog.getSaveFileName(self, self.tr("Сохранить настройки"), '', self.tr("JSON файлы (*.json)"))
            if settings_path:
                with open(settings_path, 'w', encoding='utf-8') as f:
                    json.dump(mapping, f)
                QMessageBox.information(self, self.tr("Успех"), self.tr("Настройки успешно сохранены."))
        except Exception as e:
            self.log_error(e)
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Произошла ошибка при сохранении настроек: ") + str(e))

    def get_current_mapping(self):
        if hasattr(self, 'page_match') and self.are_all_items_files(os.listdir(self.folder_path)):
            return {file: combobox.currentText() for file, combobox in self.file_to_column.items()}
        else:
            return {folder: combobox.currentText() for folder, combobox in self.folder_to_column.items()}

    def load_mapping_settings(self):
        try:
            settings_path, _ = QFileDialog.getOpenFileName(self, self.tr("Загрузить настройки"), '', self.tr("JSON файлы (*.json)"))
            if settings_path:
                with open(settings_path, 'r', encoding='utf-8') as f:
                    mapping = json.load(f)
                self.apply_loaded_mapping(mapping)
        except Exception as e:
            self.log_error(e)
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Произошла ошибка при загрузке настроек: ") + str(e))

    def apply_loaded_mapping(self, mapping):
        if hasattr(self, 'page_match') and self.are_all_items_files(os.listdir(self.folder_path)):
            self.apply_mapping_to_comboboxes(mapping, self.file_to_column)
        else:
            self.apply_mapping_to_comboboxes(mapping, self.folder_to_column)

    def apply_mapping_to_comboboxes(self, mapping, combobox_dict):
        for name, column_name in mapping.items():
            if name in combobox_dict:
                index = combobox_dict[name].findText(column_name)
                if index != -1:
                    combobox_dict[name].setCurrentIndex(index)

    # ======= Процесс копирования =======
    def start_copying(self):
        try:
            self.go_to_progress_page()
            self.perform_copying()
        except Exception as e:
            self.log_error(e)
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Произошла ошибка при запуске процесса копирования: ") + str(e))

    def perform_copying(self):
        try:
            all_successful = True
            total_items = self.calculate_total_items()
            self.progress_bar.setMaximum(total_items)
            current_progress = 0
            for sheet_name in self.selected_sheets:
                sheet = self.workbook[sheet_name]
                copy_col_index = excel_column_to_index(self.sheet_to_column[sheet_name].text())
                header_row = self.header_row[sheet_name]

                items = self.file_to_column.items() if self.file_to_column else self.folder_to_column.items()
                for name, combobox in items:
                    column_name = combobox.currentText()
                    if not column_name:
                        continue
                    if column_name not in self.columns[sheet_name]:
                        error_message = self.tr("Столбец '{0}' не найден на листе '{1}' основного файла Excel.").format(column_name, sheet_name)
                        QMessageBox.critical(self, self.tr("Ошибка"), error_message)
                        self.log_error(error_message)
                        return

                    if not self.copy_data_between_sheets(name, sheet_name, copy_col_index, header_row, column_name):
                        all_successful = False
                        return

                    current_progress += 1
                    self.progress_bar.setValue(current_progress)
                    QApplication.processEvents()

            if all_successful:
                self.finalize_copying_process()
        except Exception as e:
            self.log_error(e)
            QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Произошла ошибка при обработке файлов: ") + str(e))

    def calculate_total_items(self):
        if self.file_to_column:
            return len(self.selected_sheets) * len([c for c in self.file_to_column.values() if c.currentText()])
        elif self.folder_to_column:
            return len(self.selected_sheets) * len([c for c in self.folder_to_column.values() if c.currentText()])
        return 0

    def copy_data_between_sheets(self, name, sheet_name, copy_col_index, header_row, column_name):
        col_index = self.columns[sheet_name].index(column_name) + 1
        if self.file_to_column:
            file_path = os.path.join(self.folder_path, name)
            return self.copy_data_from_file(file_path, sheet_name, copy_col_index, header_row, col_index, name, column_name)
        else:
            lang_folder_path = os.path.join(self.folder_path, name)
            return self.copy_data_from_folder(lang_folder_path, sheet_name, copy_col_index, header_row, col_index, name, column_name)

    def copy_data_from_file(self, file_path, sheet_name, copy_col_index, header_row, col_index, name, column_name):
        if os.path.isfile(file_path) and name.endswith(('.xlsx', '.xls')):
            try:
                lang_wb = load_workbook(file_path)
                lang_sheet = lang_wb[sheet_name]
                start_row = 2 if self.skip_first_row else 1
                for row in range(start_row, lang_sheet.max_row + 1):
                    if self.copy_by_row_number and row == header_row + 1:
                        continue
                    if not self.copy_cell_value(lang_sheet, sheet_name, row, copy_col_index, header_row, col_index):
                        self.show_copy_error(name, column_name, sheet_name)
                        return False
            except Exception as e:
                self.show_file_processing_error(name, str(e))
                return False
            finally:
                if 'lang_wb' in locals():
                    lang_wb.close()
        return True

    def copy_data_from_folder(self, lang_folder_path, sheet_name, copy_col_index, header_row, col_index, name, column_name):
        for filename in os.listdir(lang_folder_path):
            file_path = os.path.join(lang_folder_path, filename)
            if os.path.isfile(file_path) and filename.endswith(('.xlsx', '.xls')):
                try:
                    lang_wb = load_workbook(file_path)
                    lang_sheet = lang_wb[sheet_name]
                    start_row = 2 if self.skip_first_row else 1
                    for row in range(start_row, lang_sheet.max_row + 1):
                        if self.copy_by_row_number and row == header_row + 1:
                            continue
                        if not self.copy_cell_value(lang_sheet, sheet_name, row, copy_col_index, header_row, col_index):
                            self.show_copy_error(filename, column_name, sheet_name)
                            lang_wb.close()
                            return False
                    lang_wb.close()
                except Exception as e:
                    self.show_file_processing_error(filename, str(e))
                    return False
        return True

    def show_copy_error(self, file_name, column_name, sheet_name):
        error_message = self.tr("Ошибка копирования значения из файла {0} в столбец {1} на листе {2}").format(file_name, column_name, sheet_name)
        QMessageBox.critical(self, self.tr("Ошибка"), error_message)
        self.log_error(error_message)

    def show_file_processing_error(self, file_name, error_message):
        error_msg = self.tr("Ошибка при обработке файла {0}: {1}").format(file_name, error_message)
        QMessageBox.critical(self, self.tr("Ошибка"), error_msg)
        self.log_error(error_msg)

    def copy_cell_value(self, lang_sheet, sheet_name, row, copy_col_index, header_row, col_index):
        source_value = lang_sheet.cell(row=row, column=copy_col_index).value
        if source_value is None:
            return True

        if self.copy_by_row_number:
            target_row = row
        else:
            start_row_offset = 2 if self.skip_first_row else 1
            target_row = header_row + start_row_offset + (row - start_row_offset)

        target_cell = self.workbook[sheet_name].cell(row=target_row, column=col_index)

        def compute_hash(text):
            if text is None:
                text = ""
            return hashlib.sha256(text.encode('utf-8')).hexdigest()

        source_hash = compute_hash(source_value)

        max_attempts = 5
        for attempt in range(max_attempts):
            target_cell.value = source_value
            if source_value == target_cell.value:
                if compute_hash(target_cell.value) == source_hash:
                    return True

        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        target_cell.fill = fill

        error_message = self.tr(
            "Несовпадение текста при копировании ячейки {0}{1} на листе {2}.\n"
            "Исходный текст:\n{3}\n\n"
            "Содержимое целевой ячейки:\n{4}\n\n"
            "Контрольные суммы: исходный {5} vs целевой {6}."
        ).format(
            utils.get_column_letter(col_index), target_row, sheet_name,
            source_value, target_cell.value,
            source_hash, compute_hash(target_cell.value)
        )

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle(self.tr("Предупреждение"))
        msg_box.setText(self.tr("Ошибка копирования ячейки"))
        msg_box.setInformativeText(error_message)
        retry_button = msg_box.addButton(self.tr("Повторить"), QMessageBox.AcceptRole)
        save_button = msg_box.addButton(self.tr("Сохранить как есть"), QMessageBox.AcceptRole)
        abort_button = msg_box.addButton(self.tr("Остановить"), QMessageBox.RejectRole)
        msg_box.exec()

        if msg_box.clickedButton() == retry_button:
            return self.copy_cell_value(lang_sheet, sheet_name, row, copy_col_index, header_row, col_index)
        elif msg_box.clickedButton() == save_button:
            return True
        else:
            return False

    def finalize_copying_process(self):
        base, ext = os.path.splitext(self.excel_file_path)
        output_file = f"{base}_out{ext}"
        self.workbook.save(output_file)
        self.workbook.close()
        self.workbook = None
        self.go_to_completion_page()