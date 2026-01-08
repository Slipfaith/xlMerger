import os

import traceback
import pandas as pd
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QComboBox, QTableView, QMessageBox,
    QHeaderView, QCheckBox, QMenu
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QStandardItemModel, QStandardItem, QColor, QBrush
from openpyxl import load_workbook
import openpyxl.utils as utils


class ClickableHeaderView(QHeaderView):
    """Header view emitting signals on left and right clicks."""

    leftClicked = Signal(int)
    rightClicked = Signal(int)

    def mousePressEvent(self, event):
        index = self.logicalIndexAt(event.position().toPoint())
        if index >= 0:
            if event.button() == Qt.LeftButton:
                self.leftClicked.emit(index)
            elif event.button() == Qt.RightButton:
                self.rightClicked.emit(index)
        super().mousePressEvent(event)

class ExcelPreviewer(QWidget):
    def __init__(self, excel_path):
        super().__init__()
        self.excel_path = None
        self.workbook = None
        self.sheet_names = []
        self.current_sheet = None
        self.source_col = None
        self.target_cols = set()
        self.init_ui()
        self.load_file(excel_path)

    def closeEvent(self, event):
        if self.workbook:
            self.workbook.close()
        super().closeEvent(event)

    def init_ui(self):
        layout = QVBoxLayout()

        self.select_all_checkbox = QCheckBox(self.tr("Выбрать все листы"))
        self.select_all_checkbox.toggled.connect(self.toggle_all_sheets)
        layout.addWidget(self.select_all_checkbox)

        self.init_sheet_selector()
        layout.addWidget(self.sheet_selector)

        self.table_view = QTableView(self)
        header = ClickableHeaderView(Qt.Horizontal, self.table_view)
        header.leftClicked.connect(self.handle_left_click)
        header.rightClicked.connect(self.handle_right_click)
        self.table_view.setHorizontalHeader(header)
        layout.addWidget(self.table_view)

        self.setLayout(layout)
        self.setGeometry(100, 100, 800, 600)
        self.show()

    def load_file(self, excel_path):
        if self.workbook:
            self.workbook.close()
        self.excel_path = excel_path
        self.workbook = load_workbook(excel_path, read_only=True)
        self.sheet_names = self.workbook.sheetnames
        self.current_sheet = self.sheet_names[0] if self.sheet_names else ""
        self.source_col = None
        self.target_cols.clear()
        self.init_sheet_selector()
        self.setWindowTitle(self.tr("Просмотр: ") + self.current_sheet)
        self.load_excel_data()

    def init_sheet_selector(self):
        if not hasattr(self, "sheet_selector"):
            self.sheet_selector = QComboBox(self)
            self.sheet_selector.currentTextChanged.connect(self.switch_sheet)
        self.sheet_selector.blockSignals(True)
        self.sheet_selector.clear()
        self.sheet_selector.addItems(self.sheet_names)
        self.sheet_selector.setEnabled(self.select_all_checkbox.isChecked())
        if self.current_sheet:
            self.sheet_selector.setCurrentText(self.current_sheet)
        self.sheet_selector.blockSignals(False)

    def toggle_all_sheets(self, checked):
        self.sheet_selector.setEnabled(checked)
        self.sheet_selector.clear()
        if checked:
            self.sheet_selector.addItems(self.sheet_names)
            if self.current_sheet not in self.sheet_names:
                self.current_sheet = self.sheet_names[0]
            self.sheet_selector.setCurrentText(self.current_sheet)
        else:
            self.sheet_selector.addItem(self.current_sheet)

    def switch_sheet(self, sheet_name):
        self.current_sheet = sheet_name
        self.source_col = None
        self.target_cols.clear()
        self.load_excel_data()

    def load_excel_data(self):
        try:
            df = self.read_excel_data()
            self.populate_table_view(df)
        except Exception as e:
            self.handle_load_error(e)

    def read_excel_data(self):
        df = pd.read_excel(self.excel_path, sheet_name=self.current_sheet, header=None, dtype=str)
        df.fillna("", inplace=True)
        return df

    def populate_table_view(self, df):
        model = self.create_table_model(df)
        self.table_view.setModel(model)
        self.adjust_table_headers(df)
        self.update_colors()

    def create_table_model(self, df):
        model = QStandardItemModel(df.shape[0], df.shape[1])
        # Буквенные заголовки столбцов (A, B, C, ...)
        model.setHorizontalHeaderLabels([utils.get_column_letter(col + 1) for col in range(df.shape[1])])
        row_labels = [str(i + 1) for i in range(df.shape[0])]
        model.setVerticalHeaderLabels(row_labels)

        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                item_value = df.iat[row, col]
                # Можно убрать или оставить print для отладки
                # print(f"Строка: {row+1}, Столбец: {col+1}, Длина исходного текста: {len(item_value)}, Текст: {item_value}")
                item = QStandardItem(item_value)
                model.setItem(row, col, item)
        return model

    def adjust_table_headers(self, df):
        vertical_header = self.table_view.verticalHeader()
        vertical_header.setVisible(True)
        vertical_header.setMinimumWidth(40)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table_view.verticalHeader().setDefaultSectionSize(20)
        self.table_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        max_width_px = 113

        for col in range(df.shape[1]):
            self.table_view.resizeColumnToContents(col)
            current_width = self.table_view.columnWidth(col)
            new_width = min(current_width, max_width_px)
            self.table_view.setColumnWidth(col, new_width)

        self.update_colors()

    def handle_load_error(self, error):
        QMessageBox.critical(self, self.tr("Ошибка"), self.tr("Не удалось загрузить данные с листа: ") + str(error))
        self.log_error(error)

    def log_error(self, error):
        log_file_path = os.path.join(os.path.dirname(__file__), 'error_log.txt')
        with open(log_file_path, 'a', encoding='utf-8') as log_file:
            log_file.write(self.tr("Ошибка: ") + str(error) + "\n")
            log_file.write(traceback.format_exc())
            log_file.write("\n")

    # --- Column selection logic ---
    def handle_left_click(self, index: int):
        if self.source_col is None:
            self.source_col = index
            self.target_cols.clear()
        else:
            if index == self.source_col or index in self.target_cols:
                return
            cols = {self.source_col} | self.target_cols
            min_idx = min(cols)
            max_idx = max(cols)
            if index == min_idx - 1 or index == max_idx + 1:
                self.target_cols.add(index)
        self.update_colors()

    def handle_right_click(self, index: int):
        if index == self.source_col:
            self.source_col = None
            self.target_cols.clear()
        elif index in self.target_cols:
            self.target_cols.remove(index)
        else:
            return
        self.update_colors()

    def update_colors(self):
        model = self.table_view.model()
        if model is None:
            return
        for row in range(model.rowCount()):
            for col in range(model.columnCount()):
                idx = model.index(row, col)
                model.setData(idx, QBrush(QColor("white")), Qt.BackgroundRole)
        if self.source_col is not None:
            for row in range(model.rowCount()):
                idx = model.index(row, self.source_col)
                model.setData(idx, QBrush(QColor("#a2cffe")), Qt.BackgroundRole)
        for col in self.target_cols:
            for row in range(model.rowCount()):
                idx = model.index(row, col)
                model.setData(idx, QBrush(QColor("#b6fcb6")), Qt.BackgroundRole)
