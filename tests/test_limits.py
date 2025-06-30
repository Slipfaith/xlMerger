import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from core.limit_auto import check_limits_auto
from core.limit_manual import check_limits_manual
from utils.utils import excel_column_to_index


def test_check_limits_auto_basic():
    wb = Workbook()
    ws = wb.active
    ws.append(["Limit", "Text1", "Text2"])
    ws.append([5, "hello", "toolong"])
    ws.append([5, "ok", "short"])
    headers = ["Limit", "Text1", "Text2"]
    mappings = [
        ("Limit", ["Text2"], False, None, None, "column")
    ]

    report, total = check_limits_auto(ws, headers, mappings)
    assert total == 1
    assert report
    fill = ws.cell(row=2, column=3).fill
    assert fill.fill_type == "solid"


def test_check_limits_auto_manual_limits():
    wb = Workbook()
    ws = wb.active
    ws.append(["L", "T"])
    ws.append([None, "abc"])
    headers = ["L", "T"]
    mappings = [
        ("L", ["T"], True, 2, None, "column")
    ]

    report, total = check_limits_auto(ws, headers, mappings)
    assert total == 1
    assert "лимит" in report[0]


def test_check_limits_manual_cells():
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["longtext", "b"])
    headers = ["A", "B"]
    mappings = [
        ([(0, 0)], True, 3, None, "cell")
    ]

    report, total = check_limits_manual(ws, headers, mappings)
    assert total == 1
    assert ws.cell(row=2, column=1).fill.fill_type == "solid"


def test_excel_column_to_index():
    assert excel_column_to_index("A") == 1
    assert excel_column_to_index("Z") == 26
    assert excel_column_to_index("AA") == 27
