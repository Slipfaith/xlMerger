# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from core.merge_columns import merge_excel_columns, _get_data_max_row


def create_wb(path, data, sheet_name="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, values in data.items():
        for idx, val in enumerate(values, start=1):
            ws[f"{col}{idx}"] = val
    wb.save(path)
    wb.close()


def test_merge_excel_columns(tmp_path):
    main_path = tmp_path / "main.xlsx"
    src_path = tmp_path / "src.xlsx"

    create_wb(main_path, {"A": ["h1", "h2", "h3"]}, sheet_name="Main")
    create_wb(src_path, {"A": ["x", "y", "z"]})

    mappings = [{
        "source": str(src_path),
        "source_columns": ["A"],
        "target_sheet": "Main",
        "target_columns": ["B"],
    }]

    output = merge_excel_columns(str(main_path), mappings)
    wb = load_workbook(output)
    ws = wb["Main"]
    result = [ws[f"B{i}"].value for i in range(1, 4)]
    assert result == ["x", "y", "z"]
    wb.close()


def test_merge_excel_columns_skips_empty_rows(tmp_path):
    main_path = tmp_path / "main.xlsx"
    src_path = tmp_path / "src.xlsx"

    create_wb(main_path, {}, sheet_name="Main")
    values = [None, "x", None, None, "y", None, "z", "a", "b", "c"]
    create_wb(src_path, {"A": values})

    mappings = [{
        "source": str(src_path),
        "source_columns": ["A"],
        "target_sheet": "Main",
        "target_columns": ["B"],
    }]

    output = merge_excel_columns(str(main_path), mappings)
    wb = load_workbook(output)
    ws = wb["Main"]

    assert ws["B2"].value == "x"
    assert ws["B5"].value == "y"
    assert ws["B7"].value == "z"
    assert ws["B8"].value == "a"
    assert ws["B9"].value == "b"
    assert ws["B10"].value == "c"

    assert ws["B1"].value is None
    assert ws["B3"].value is None
    assert ws["B4"].value is None
    assert ws["B6"].value is None
    wb.close()


def test_get_data_max_row_ignores_style_only_rows(tmp_path):
    src_path = tmp_path / "style_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A5000"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    wb.save(src_path)
    wb.close()

    wb = load_workbook(src_path, data_only=True)
    ws = wb.active
    assert _get_data_max_row(ws) == 0
    wb.close()


def test_merge_excel_columns_handles_style_only_source_sheet(tmp_path):
    main_path = tmp_path / "main.xlsx"
    src_path = tmp_path / "src_style_only.xlsx"

    create_wb(main_path, {"A": ["h1", "h2", "h3"]}, sheet_name="Main")
    wb_src = Workbook()
    ws_src = wb_src.active
    ws_src["A7000"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    wb_src.save(src_path)
    wb_src.close()

    mappings = [{
        "source": str(src_path),
        "source_columns": ["A"],
        "target_sheet": "Main",
        "target_columns": ["B"],
    }]

    output = merge_excel_columns(str(main_path), mappings)
    wb = load_workbook(output)
    ws = wb["Main"]
    assert ws["B1"].value is None
    assert ws["B2"].value is None
    assert ws["B3"].value is None
    wb.close()
