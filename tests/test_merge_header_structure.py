# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from gui.merge_mapping_dialog import get_excel_structure
from gui.multi_merge_mapping_dialog import _read_structure


def _create_inflated_header_book(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "RU"
    ws["B1"] = "KA"
    # inflate max_column without adding real header values
    ws["XFD1"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    wb.save(path)
    wb.close()


def test_get_excel_structure_trims_trailing_empty_headers(tmp_path):
    path = tmp_path / "inflated.xlsx"
    _create_inflated_header_book(path)

    structure = get_excel_structure(str(path))
    assert structure["Main"] == ["RU", "KA"]


def test_read_structure_trims_trailing_empty_headers(tmp_path):
    path = tmp_path / "inflated.xlsx"
    _create_inflated_header_book(path)

    structure = _read_structure(str(path))
    assert structure["Main"]["headers"] == ["RU", "KA"]
