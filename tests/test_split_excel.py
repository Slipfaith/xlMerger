import os
from openpyxl import Workbook, load_workbook
from core.split_excel import split_excel_by_languages, split_excel_multiple_sheets


def test_split_excel_without_header(tmp_path):
    src = tmp_path / "main.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([None, None])
    ws.append(["привет", "hallo"])
    wb.save(src)
    wb.close()

    split_excel_by_languages(str(src), "Sheet1", "A", target_langs=["B"])

    out_file = tmp_path / "main_A-B.xlsx"
    assert out_file.is_file()
    wb_out = load_workbook(out_file)
    ws_out = wb_out.active
    assert ws_out.cell(row=1, column=1).value == "A"
    assert ws_out.cell(row=1, column=2).value == "B"
    assert ws_out.cell(row=2, column=1).value == "привет"
    assert ws_out.cell(row=2, column=2).value == "hallo"
    wb_out.close()

def test_split_excel(tmp_path):
    src = tmp_path / "main.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ru", "de", "en", "comment"])
    ws.append(["привет", "hallo", "hello", "c1"])
    wb.save(src)
    wb.close()

    split_excel_by_languages(str(src), "Sheet1", "ru")

    out_de = tmp_path / "main_ru-de.xlsx"
    out_en = tmp_path / "main_ru-en.xlsx"

    assert out_de.is_file()
    assert out_en.is_file()

    wb_de = load_workbook(out_de)
    ws_de = wb_de.active
    assert ws_de.cell(row=1, column=1).value == "ru"
    assert ws_de.cell(row=1, column=2).value == "de"
    assert ws_de.cell(row=2, column=1).value == "привет"
    assert ws_de.cell(row=2, column=2).value == "hallo"
    wb_de.close()
    wb_en = load_workbook(out_en)
    ws_en = wb_en.active
    assert ws_en.cell(row=1, column=1).value == "ru"
    assert ws_en.cell(row=1, column=2).value == "en"
    assert ws_en.cell(row=2, column=1).value == "привет"
    assert ws_en.cell(row=2, column=2).value == "hello"
    wb_en.close()

def test_split_excel_with_targets(tmp_path):
    src = tmp_path / "main.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ru", "de", "en", "comment"])
    ws.append(["hi", "de_val", "en_val", "c1"])
    wb.save(src)
    wb.close()

    split_excel_by_languages(str(src), "Sheet1", "ru", target_langs=["de"])

    out_de = tmp_path / "main_ru-de.xlsx"
    out_en = tmp_path / "main_ru-en.xlsx"

    assert out_de.is_file()
    assert not out_en.exists()

def test_split_excel_with_non_language_target(tmp_path):
    src = tmp_path / "main.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ru", "English Text", "Another-Lang", "comment"])
    ws.append(["hi", "en_val", "other_val", "c1"])
    wb.save(src)
    wb.close()

    split_excel_by_languages(
        str(src),
        "Sheet1",
        "ru",
        target_langs=["English Text"]
    )

    out_custom = tmp_path / "main_ru-English Text.xlsx"
    assert out_custom.is_file()

def test_split_excel_multiple_sheets(tmp_path):
    src = tmp_path / "main.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "S1"
    ws1.append(["ru", "de"])
    ws1.append(["a", "b"])
    ws2 = wb.create_sheet("S2")
    ws2.append(["ru", "de"])
    ws2.append(["c", "d"])
    wb.save(src)
    wb.close()

    cfg = {
        "S1": ("ru", ["de"], []),
        "S2": ("ru", ["de"], []),
    }

    split_excel_multiple_sheets(str(src), cfg)

    out_file = tmp_path / "main_ru-de.xlsx"
    assert out_file.is_file()
    out_wb = load_workbook(out_file)
    assert set(out_wb.sheetnames) == {"S1", "S2"}
    out_wb.close()

def test_split_preserves_format(tmp_path):
    src = tmp_path / "main.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    from openpyxl.styles import Font, PatternFill
    ws.append(["ru", "de"])
    ws.append(["a", "b"])
    ws.cell(row=2, column=1).font = Font(bold=True)
    ws.cell(row=2, column=1).fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    wb.save(src)
    wb.close()

    split_excel_by_languages(str(src), "Sheet1", "ru")

    out_file = tmp_path / "main_ru-de.xlsx"
    wb2 = load_workbook(out_file)
    ws2 = wb2.active
    assert ws2.cell(row=2, column=1).font.bold is True

    # Исправленная проверка цвета
    fill = ws2.cell(row=2, column=1).fill.fgColor.rgb
    assert fill is not None
    assert fill.upper().endswith("FFFF00")
    wb2.close()

def test_split_ignores_empty_trailing_rows(tmp_path):
    src = tmp_path / "main.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ru", "en"])
    ws.append(["hi", "hello"])
    ws.row_dimensions[100].height = 15
    wb.save(src)
    wb.close()

    split_excel_by_languages(str(src), "Sheet1", "ru")

    out_file = tmp_path / "main_ru-en.xlsx"
    wb2 = load_workbook(out_file)
    ws2 = wb2.active
    assert ws2.max_row == 2
    wb2.close()
