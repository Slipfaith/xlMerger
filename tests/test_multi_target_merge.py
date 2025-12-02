import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.multi_target_merge import (
    TargetMergePlan,
    merge_translations_to_many_targets,
    normalize_name,
    suggest_target_mapping,
)


def _create_workbook(path: Path, sheet_name: str, value: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = "header"
    ws["B1"] = "text"
    ws["B2"] = value
    wb.save(path)


def test_normalize_name_strips_special_chars():
    assert normalize_name("/tmp/My File_v1.xlsx") == "myfilev1"


def test_suggest_target_mapping_prefers_name_similarity(tmp_path):
    targets = [tmp_path / "core_project.xlsx", tmp_path / "other.xlsx"]
    sources = [tmp_path / "core_project_ru.xlsx", tmp_path / "misc.xlsx"]

    mapping = suggest_target_mapping([str(s) for s in sources], [str(t) for t in targets])

    assert mapping[str(targets[0])] == [str(sources[0])]
    assert str(sources[1]) in mapping[str(targets[1])]


def test_manual_mapping_overrides_auto(tmp_path):
    targets = [tmp_path / "alpha.xlsx", tmp_path / "beta.xlsx"]
    sources = [tmp_path / "alpha_ru.xlsx", tmp_path / "beta_ru.xlsx"]
    manual = {str(sources[0]): str(targets[1])}

    mapping = suggest_target_mapping([str(s) for s in sources], [str(t) for t in targets], manual)

    assert mapping[str(targets[1])][0] == str(sources[0])


def test_merge_translations_to_many_targets(tmp_path):
    target_a = tmp_path / "target_a.xlsx"
    target_b = tmp_path / "target_b.xlsx"
    _create_workbook(target_a, "Sheet1", "placeholder")
    _create_workbook(target_b, "Sheet1", "placeholder")

    src_a = tmp_path / "A_project_ru.xlsx"
    src_b = tmp_path / "B_project_ru.xlsx"
    _create_workbook(src_a, "Sheet1", "Translated A")
    _create_workbook(src_b, "Sheet1", "Translated B")

    plan_a = TargetMergePlan(
        main_file=str(target_a),
        mappings=[{
            "source_id": "A_project",
            "source_columns": ["B"],
            "target_sheet": "Sheet1",
            "target_columns": ["B"],
        }],
        output_file=str(tmp_path / "target_a_out.xlsx"),
    )
    plan_b = TargetMergePlan(
        main_file=str(target_b),
        mappings=[{
            "source_id": "B_project",
            "source_columns": ["B"],
            "target_sheet": "Sheet1",
            "target_columns": ["B"],
        }],
        output_file=str(tmp_path / "target_b_out.xlsx"),
    )

    outputs = merge_translations_to_many_targets([plan_a, plan_b], [str(src_a), str(src_b)])

    assert set(outputs) == {plan_a.output_file, plan_b.output_file}
    for output_file, expected in zip(outputs, ("Translated A", "Translated B")):
        wb = load_workbook(output_file)
        ws = wb["Sheet1"]
        assert ws["B2"].value == expected
        wb.close()
