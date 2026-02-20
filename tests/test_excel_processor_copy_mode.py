# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from core.excel_processor import ExcelProcessor


class _DummyLogger:
    def log_error(self, *args, **kwargs):
        return None

    def log_info(self, *args, **kwargs):
        return None

    def log_copy(self, *args, **kwargs):
        return None

    def save(self):
        return None


def _create_main_book(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Target"])
    ws.append(["k1", None])
    ws.append(["k2", None])
    ws.append(["k3", None])
    wb.save(path)
    wb.close()


def _create_source_book(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["SRC"])
    ws.append(["v1"])
    ws.append([""])
    ws.append(["v3"])
    wb.save(path)
    wb.close()


def _run_copy(tmp_path, copy_by_row_number):
    main = tmp_path / ("main_row.xlsx" if copy_by_row_number else "main_seq.xlsx")
    src = tmp_path / ("src_row.xlsx" if copy_by_row_number else "src_seq.xlsx")
    _create_main_book(main)
    _create_source_book(src)

    processor = ExcelProcessor(
        main_excel_path=str(main),
        folder_path="",
        copy_column="A",
        selected_sheets=["Sheet1"],
        sheet_to_header_row={"Sheet1": 0},
        sheet_to_column={"Sheet1": "A"},
        file_to_column={str(src): "Target"},
        folder_to_column={},
        file_to_sheet_map={},
        skip_first_row=True,
        copy_by_row_number=copy_by_row_number,
        preserve_formatting=False,
        logger=_DummyLogger(),
    )
    output = processor.copy_data()
    wb = load_workbook(output)
    ws = wb["Sheet1"]
    values = [ws.cell(row=2, column=2).value, ws.cell(row=3, column=2).value, ws.cell(row=4, column=2).value]
    wb.close()
    return values


def test_copy_data_sequential_mode_compacts_non_empty_rows(tmp_path):
    values = _run_copy(tmp_path, copy_by_row_number=False)
    assert values == ["v1", "v3", None]


def test_copy_data_row_number_mode_preserves_source_row_indexes(tmp_path):
    values = _run_copy(tmp_path, copy_by_row_number=True)
    assert values == ["v1", None, "v3"]


def test_copy_data_row_number_mode_without_skip_keeps_absolute_row_numbers(tmp_path):
    main = tmp_path / "main_abs.xlsx"
    src = tmp_path / "src_abs.xlsx"
    _create_main_book(main)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["v1"])
    ws.append([""])
    ws.append(["v3"])
    wb.save(src)
    wb.close()

    processor = ExcelProcessor(
        main_excel_path=str(main),
        folder_path="",
        copy_column="A",
        selected_sheets=["Sheet1"],
        sheet_to_header_row={"Sheet1": 0},
        sheet_to_column={"Sheet1": "A"},
        file_to_column={str(src): "Target"},
        folder_to_column={},
        file_to_sheet_map={},
        skip_first_row=False,
        copy_by_row_number=True,
        preserve_formatting=False,
        logger=_DummyLogger(),
    )
    output = processor.copy_data()
    wb = load_workbook(output)
    ws = wb["Sheet1"]
    assert ws.cell(row=1, column=2).value == "v1"
    assert ws.cell(row=2, column=2).value is None
    assert ws.cell(row=3, column=2).value == "v3"
    wb.close()


def test_get_data_max_row_ignores_style_only_rows(tmp_path):
    src = tmp_path / "style_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A12000"].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    wb.save(src)
    wb.close()

    wb = load_workbook(src, data_only=True)
    ws = wb.active
    assert ExcelProcessor._get_data_max_row(ws) == 0
    wb.close()
