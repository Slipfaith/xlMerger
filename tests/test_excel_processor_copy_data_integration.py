# core/excel_processor.py

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import os

class ExcelProcessor:
    def __init__(
        self,
        main_excel_path,
        folder_path,
        copy_column="A",
        selected_sheets=None,
        sheet_to_header_row=None,
        sheet_to_column=None,
        folder_to_column=None,
        logger=None
    ):
        self.main_excel_path = main_excel_path
        self.folder_path = folder_path
        self.copy_column = copy_column
        self.selected_sheets = selected_sheets or []
        self.sheet_to_header_row = sheet_to_header_row or {}
        self.sheet_to_column = sheet_to_column or {}
        self.folder_to_column = folder_to_column or {}
        self.logger = logger

    def validate_paths_and_column(self):
        if not os.path.isfile(self.main_excel_path):
            raise FileNotFoundError(f"Файл не найден: {self.main_excel_path}")
        if not os.path.isdir(self.folder_path):
            raise FileNotFoundError(f"Папка не найдена: {self.folder_path}")

        for sheet in self.selected_sheets:
            col = self.sheet_to_column.get(sheet)
            if not col:
                raise ValueError(f"Для листа '{sheet}' не указан столбец для копирования")
        # Валидация прошла — можно логировать
        if self.logger:
            self.logger.info(f"Валидация путей и колонок прошла успешно.")

    @staticmethod
    def column_letter_by_header(ws, header, header_row=0):
        """Вернуть букву столбца по имени заголовка."""
        for col_idx, cell in enumerate(ws[header_row+1], 1):
            if str(cell.value).strip() == header:
                return get_column_letter(col_idx)
        raise ValueError(f"Заголовок '{header}' не найден в строке {header_row+1}")

    def copy_data(self):
        wb_main = load_workbook(self.main_excel_path)
        if self.logger:
            self.logger.info(f"Загружен основной Excel: {self.main_excel_path}")

        for sheet_name in self.selected_sheets:
            ws_main = wb_main[sheet_name]
            if self.logger:
                self.logger.info(f"Обрабатывается лист: {sheet_name}")

            # Определяем букву столбца для копирования
            col_key = self.sheet_to_column[sheet_name]
            header_row = self.sheet_to_header_row.get(sheet_name, 0)

            # Определяем букву колонки: если имя заголовка — ищем, если буква — используем сразу
            if not col_key.isalpha():
                col_letter = self.column_letter_by_header(ws_main, col_key, header_row)
            else:
                col_letter = col_key

            col_idx = column_index_from_string(col_letter)

            # Копируем данные из переводных файлов
            folder_name = os.path.basename(self.folder_path)
            folder_col_key = self.folder_to_column.get(folder_name, col_key)
            # Если для папки задано имя заголовка, найдём нужную букву
            if not folder_col_key.isalpha():
                folder_col_letter = self.column_letter_by_header(ws_main, folder_col_key, header_row)
            else:
                folder_col_letter = folder_col_key
            folder_col_idx = column_index_from_string(folder_col_letter)

            for fname in os.listdir(self.folder_path):
                if not fname.endswith(".xlsx"):
                    continue
                trans_path = os.path.join(self.folder_path, fname)
                wb_trans = load_workbook(trans_path)
                ws_trans = wb_trans[sheet_name]
                # Копируем значения (строки после заголовка)
                for row_idx in range(header_row+2, ws_main.max_row+1):
                    main_val = ws_main.cell(row=row_idx, column=col_idx).value
                    trans_val = ws_trans.cell(row=row_idx, column=folder_col_idx).value
                    # Копируем только если в переводе что-то есть
                    if trans_val is not None and trans_val != "":
                        ws_main.cell(row=row_idx, column=col_idx).value = trans_val

                wb_trans.close()

        # Сохраняем новый файл с суффиксом
        base, ext = os.path.splitext(self.main_excel_path)
        output_file = base + "_copydata" + ext
        wb_main.save(output_file)
        wb_main.close()
        if self.logger:
            self.logger.info(f"Файл успешно сохранён: {output_file}")
        return output_file
